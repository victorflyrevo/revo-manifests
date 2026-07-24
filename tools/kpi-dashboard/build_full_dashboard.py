#!/usr/bin/env python3
"""Build local dashboard with longest history + MoM / YoY growth charts.

Reads boardings from Manifests API export (or local CSV) — outside git.
"""

from __future__ import annotations

import csv
import json
import os
import re
import sys
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Optional
from urllib.request import Request, urlopen

OUT = Path(__file__).resolve().parent

# Reuse API mission cut (Sigtrip-style connected chains)
_REPO_ROOT = OUT.parents[1]  # tools/kpi-dashboard → repo root
for candidate in (_REPO_ROOT, Path("/workspace")):
    if (candidate / "app" / "missions.py").exists():
        if str(candidate) not in sys.path:
            sys.path.insert(0, str(candidate))
        break
from app.missions import MissionLeg, assign_missions, missions_by_month  # noqa: E402

# Match Cancelado / CANCELADO / CANCELAD (truncated) anywhere in the sheet tab
_CANCELLED_SHEET = re.compile(r"\bcancel", re.I)


def is_cancelled_sheet(sheet_name: str) -> bool:
    return bool(_CANCELLED_SHEET.search(sheet_name or ""))


def is_siav_loop(row: dict) -> bool:
    o = str(row.get("origin_code") or "").strip().upper()
    d = str(row.get("dest_code") or "").strip().upper()
    return o == "SIAV" and d == "SIAV"


def drop_cancelled_rows(rows: list[dict]) -> list[dict]:
    kept = []
    removed_cancel = removed_siav = 0
    for r in rows:
        if is_cancelled_sheet(str(r.get("sheet_name") or "")):
            removed_cancel += 1
            continue
        if is_siav_loop(r):
            removed_siav += 1
            continue
        kept.append(r)
    if removed_cancel:
        print(f"Dropped {removed_cancel} cancelled-sheet boarding row(s)")
    if removed_siav:
        print(f"Dropped {removed_siav} SIAV→SIAV training boarding row(s)")
    return kept


def month_start(d: date) -> date:
    return d.replace(day=1)


def add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    return date(y, m, min(d.day, monthrange(y, m)[1]))


def month_end(d: date) -> date:
    return add_months(month_start(d), 1) - timedelta(days=1)


def month_label(d: date) -> str:
    return d.strftime("%Y-%m")


def iter_months(start: date, end: date) -> list[date]:
    cur, last = month_start(start), month_start(end)
    out = []
    while cur <= last:
        out.append(cur)
        cur = add_months(cur, 1)
    return out


def pct_change(curr: float, prev: float) -> Optional[float]:
    if prev is None or prev == 0:
        return None
    return round((curr - prev) / prev * 100, 1)


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = OUT / ".env.local"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env.setdefault(k.strip(), v.strip())
    return env


def fetch_api(base: str, api_key: str) -> tuple[str, dict, list, list, list]:
    headers = {"Accept": "application/json", "X-API-Key": api_key}

    def get_json(path: str):
        req = Request(base.rstrip("/") + path, headers=headers)
        with urlopen(req, timeout=120) as res:
            return json.loads(res.read().decode("utf-8"))

    def get_text(path: str) -> str:
        req = Request(
            base.rstrip("/") + path,
            headers={**headers, "Accept": "text/csv,*/*"},
        )
        with urlopen(req, timeout=180) as res:
            return res.read().decode("utf-8")

    summary = get_json("/api/v1/summary?days=3650")
    monthly = get_json("/api/v1/monthly")
    routes = get_json("/api/v1/routes?days=3650&limit=25")
    top = get_json("/api/v1/passengers/top?days=3650&limit=50")
    csv_text = get_text("/api/v1/export/boardings.csv")
    return csv_text, summary, monthly, routes, top


# Longest available operational window (includes 2024 uploads)
BASE_START = date(2024, 1, 1)

# Fixed reference cuts for the recurrence report
SNAPSHOT_SPECS = (
    ("2026-06", "Jun/2026 (atual)"),
    ("2025-12", "Dez/2025"),
    ("2024-12", "Dez/2024"),
)

# Exact boarding counts 1..20, plus a single >20 bucket
FREQ_CAP = 20
FREQ_KEYS = [f"ltm_freq_{i}" for i in range(1, FREQ_CAP + 1)] + ["ltm_freq_gt20"]
FREQ_LABELS = [f"{i}×" for i in range(1, FREQ_CAP + 1)] + [f">{FREQ_CAP}"]

# Shared glossary for HTML + Excel (term, definition)
GLOSSARY = (
    (
        "Unique customers",
        "Passageiros distintos na base (identidade consolidada). Conta cada pessoa uma vez no período considerado.",
    ),
    (
        "Customers (boardings)",
        "Total de embarques: cada vez que um passageiro aparece em um voo conta 1. Uma pessoa que voou 3 vezes = 3 customers/boardings.",
    ),
    (
        "Unique mês",
        "Passageiros distintos que voaram no mês civil (1º ao último dia do mês).",
    ),
    (
        "LTM (Last Twelve Months)",
        "Janela móvel de até 12 meses terminando no mês de referência (ex.: LTM de jun/2026 = jul/2025 → jun/2026).",
    ),
    (
        "Unique LTM",
        "Passageiros com pelo menos 1 boarding dentro da janela LTM.",
    ),
    (
        "Frequência (1× … 20×, >20)",
        "Quantos unique passageiros tiveram exatamente N boardings no recorte (LTM, mês-fim ou todo o período). >20 agrupa quem voou 21 vezes ou mais.",
    ),
    (
        "≥2 / recorrentes",
        "Passageiros com 2 ou mais boardings no recorte. É o corte principal de recorrência.",
    ),
    (
        "≥4",
        "Passageiros com 4 ou mais boardings no recorte — recorrência mais intensa.",
    ),
    (
        "Δ vs −12m",
        "Diferença absoluta entre o valor do LTM atual e o LTM que terminava no mesmo mês do ano anterior (ex.: jun/2026 vs jun/2025).",
    ),
    (
        "Todo o período",
        "Base completa disponível nos dados (em geral jan/2024 → último boarding), sem limitar a 12 meses.",
    ),
    (
        "Snapshots",
        "Cortes fixos de referência: Jun/2026, Dez/2025 e Dez/2024 — cada um com seu LTM.",
    ),
    (
        "Novos clientes",
        "Passageiros cuja primeira aparição na base cai naquele mês civil.",
    ),
    (
        "Cumulativo unique",
        "Soma progressiva de unique customers desde o início da base até o mês.",
    ),
    (
        "Missão (Sigtrip)",
        "Cadeia de pernas conectadas no mesmo dia / aeronave (corte operacional). Distinto de boarding de passageiro.",
    ),
    (
        "Perna (leg)",
        "Um trecho/voo individual no manifesto (origem → destino).",
    ),
    (
        "SIAV→SIAV / cancelados",
        "Treinos SIAV→SIAV e abas canceladas são excluídos das contagens deste dashboard.",
    ),
    (
        "Horas voadas (SF)",
        "Salesforce Voo__c.TempoMissao__c em minutos ÷ 60, só Status=Executado. Fonte: pull_salesforce_kpis.py.",
    ),
    (
        "SBGR vs Resto (horas)",
        "Missão inteira conta em SBGR se algum trecho toca SBGR/Guarulhos; senão vai para Resto.",
    ),
    (
        "Shuttle (horas)",
        "Voos Salesforce com Tipo contendo Shuttle (ShuttleSeat / ShuttleFullCabin).",
    ),
    (
        "Base de vendas (faturamento)",
        "Soma de Servico.ValorPago__c via Empenho no mês do voo "
        "(relatório Financeiro · Validação Weekly KPI).",
    ),
    (
        "Corporate mobility %",
        "% do faturamento cuja Conta Faturamento (fallback Comprador) "
        "é Cliente - Pessoa Jurídica.",
    ),
)


def month_minus_years(label: str, years: int = 1) -> str:
    y, m = label.split("-")
    return f"{int(y) - years}-{m}"


def empty_freq_fields(value: Any = None) -> dict[str, Any]:
    return {k: value for k in FREQ_KEYS}


def freq_rows_from_fields(fields: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for key, label in zip(FREQ_KEYS, FREQ_LABELS):
        rows.append({"key": key, "label": label, "count": fields.get(key)})
    return rows


def ltm_freq_metrics(ltm_counts: dict[int, int]) -> dict[str, Any]:
    """Passenger frequency distribution inside an LTM boarding window."""
    buckets = {i: 0 for i in range(1, FREQ_CAP + 1)}
    gt20 = 0
    ge2 = ge4 = 0
    for n in ltm_counts.values():
        if n >= 2:
            ge2 += 1
        if n >= 4:
            ge4 += 1
        if 1 <= n <= FREQ_CAP:
            buckets[n] += 1
        elif n > FREQ_CAP:
            gt20 += 1
    unique = len(ltm_counts)
    ge2_pct = round(ge2 / unique * 100, 1) if unique else 0.0
    ge4_pct = round(ge4 / unique * 100, 1) if unique else 0.0
    freq_fields = {f"ltm_freq_{i}": buckets[i] for i in range(1, FREQ_CAP + 1)}
    freq_fields["ltm_freq_gt20"] = gt20
    return {
        "ltm_unique_customers": unique,
        "ltm_repeat_customers": ge2,
        "ltm_ge2": ge2,
        "ltm_ge4": ge4,
        "ltm_ge2_pct": ge2_pct,
        "ltm_ge4_pct": ge4_pct,
        "repeat_rate_pct": ge2_pct,
        **freq_fields,
        "ltm_freq_rows": freq_rows_from_fields(freq_fields),
    }


def _delta(curr: Optional[int], prev: Optional[int]) -> Optional[int]:
    if curr is None or prev is None:
        return None
    return curr - prev


def _mission_legs_from_rows(payload_rows: list[dict], base_start: date) -> list[MissionLeg]:
    """One MissionLeg per distinct flight_id in the filtered window."""
    by_fid: dict[str, dict] = {}
    for r in payload_rows:
        fd_raw = (r.get("flight_date") or "").strip()
        if not fd_raw:
            continue
        fd = date.fromisoformat(fd_raw[:10])
        if fd < base_start:
            continue
        fid = str(r.get("flight_id") or "").strip()
        if not fid:
            continue
        by_fid.setdefault(
            fid,
            {
                "id": int(fid) if fid.isdigit() else abs(hash(fid)) % (10**9),
                "flight_date": fd,
                "flight_time": r.get("flight_time"),
                "origin_code": r.get("origin_code"),
                "dest_code": r.get("dest_code"),
                "sheet_name": r.get("sheet_name"),
                "aircraft_reg": r.get("aircraft_reg"),
            },
        )
    return [
        MissionLeg(
            flight_id=v["id"],
            flight_date=v["flight_date"],
            flight_time=v.get("flight_time"),
            origin_code=v.get("origin_code"),
            dest_code=v.get("dest_code"),
            sheet_name=v.get("sheet_name"),
            aircraft_reg=v.get("aircraft_reg"),
        )
        for v in by_fid.values()
    ]


def compute(
    payload_rows: list[dict],
    summary: dict,
    monthly_api: list,
    routes: list,
    top: list,
    source: str,
    base_start: date = BASE_START,
) -> dict:
    boardings: list[tuple[int, date]] = []
    flight_leg_months: dict[str, set] = defaultdict(set)
    pax_dates: dict[int, list[date]] = defaultdict(list)

    for r in payload_rows:
        fd_raw = (r.get("flight_date") or "").strip()
        if not fd_raw or not r.get("passenger_id"):
            continue
        fd = date.fromisoformat(fd_raw[:10])
        if fd < base_start:
            continue
        pid = int(r["passenger_id"])
        boardings.append((pid, fd))
        pax_dates[pid].append(fd)
        fl = r.get("flight_id") or f"{fd}|{r.get('sheet_name')}|{r.get('flight_time')}"
        flight_leg_months[month_label(fd)].add(str(fl))

    # Sigtrip-style missions (connected same-day chains per aircraft)
    mission_month_counts = missions_by_month(
        _mission_legs_from_rows(payload_rows, base_start)
    )

    if not boardings:
        raise SystemExit("No boardings found")

    first_seen: dict[int, date] = {}
    for pid, fd in boardings:
        if pid not in first_seen or fd < first_seen[pid]:
            first_seen[pid] = fd

    data_start = min(fd for _, fd in boardings)
    # Anchor series at requested base month even if first boarding is later
    series_start = min(month_start(data_start), month_start(base_start))
    if series_start < month_start(base_start):
        series_start = month_start(base_start)
    data_end = max(fd for _, fd in boardings)
    months_all = iter_months(month_start(base_start), data_end)

    # Fill continuous month series (zeros for gaps)
    api_by = {
        r["month"]: r
        for r in monthly_api
        if r.get("month", "") >= month_label(base_start)
    }
    monthly: list[dict] = []
    cumulative = 0
    prev_row: Optional[dict] = None
    prev_active: Optional[dict] = None
    yoy_index: dict[str, dict] = {}

    for m0 in months_all:
        label = month_label(m0)
        m_end = month_end(m0)
        new_customers = sum(1 for fd in first_seen.values() if month_start(fd) == m0)
        cumulative += new_customers

        # Active unique in calendar month
        active = {
            pid
            for pid, dates in pax_dates.items()
            if any(month_start(d) == m0 for d in dates)
        }
        boardings_m = sum(1 for _, fd in boardings if month_start(fd) == m0)
        legs_m = len(flight_leg_months.get(label, set()))
        flights_m = mission_month_counts.get(label, 0)
        if label in api_by:
            boardings_m = api_by[label]["boardings"]
            # Prefer API missions when present; else local mission cut
            if api_by[label].get("missions") is not None:
                flights_m = api_by[label]["missions"]
            elif api_by[label].get("flight_count_unit") == "mission":
                flights_m = api_by[label]["flights"]
            else:
                flights_m = mission_month_counts.get(label, api_by[label]["flights"])
            legs_m = api_by[label].get("flight_legs", legs_m)
            unique_m = api_by[label]["unique_passengers"]
        else:
            unique_m = len(active)

        has_activity = boardings_m > 0
        # Known ingest gaps / thin months (re-check after each API refresh).
        # Only flag a month when the note still matches reality (e.g. empty June).
        KNOWN_GAPS = {
            "2026-06": (
                "Mai-Jun_2026 incompleto na API (sem boardings em junho)",
                lambda n: n == 0,
            ),
        }
        data_gap = None
        if label in KNOWN_GAPS:
            note, still_broken = KNOWN_GAPS[label]
            if still_broken(boardings_m):
                data_gap = note
        if has_activity and boardings_m < 20:
            data_gap = (data_gap + " · " if data_gap else "") + f"mês com poucos boardings ({boardings_m})"

        # Rolling LTM ending this month (longest available window up to 12m)
        win_start = add_months(m0, -11)
        if win_start < month_start(data_start):
            win_start = month_start(data_start)
        ltm_counts: dict[int, int] = defaultdict(int)
        for pid, dates in pax_dates.items():
            n = sum(1 for d in dates if win_start <= d <= m_end)
            if n:
                ltm_counts[pid] = n
        freq = ltm_freq_metrics(ltm_counts)

        row = {
            "month": label,
            "new_customers": new_customers,
            "cumulative_unique_customers": cumulative,
            "unique_passengers": unique_m,
            "boardings": boardings_m,
            "flights": flights_m,
            "missions": flights_m,
            "flight_legs": legs_m,
            "has_activity": has_activity,
            "data_gap": data_gap,
            **freq,
            "window_start": win_start.isoformat(),
            "window_end": m_end.isoformat(),
            "mom_new_pct": None,
            "mom_unique_pct": None,
            "mom_boardings_pct": None,
            "mom_cumulative_pct": None,
            "yoy_new_pct": None,
            "yoy_unique_pct": None,
            "yoy_boardings_pct": None,
            "yoy_cumulative_pct": None,
            "mom_vs_month": None,
            "yoy_vs_month": None,
            "ltm_ge2_delta_vs_12m": None,
            "ltm_ge4_delta_vs_12m": None,
            "ltm_unique_delta_vs_12m": None,
            "ltm_vs_month": None,
        }

        # MoM / YoY on CUMULATIVE unique (primary growth view)
        if prev_row is not None:
            row["mom_cumulative_pct"] = pct_change(
                cumulative, prev_row["cumulative_unique_customers"]
            )
            row["mom_vs_month"] = prev_row["month"]
            # secondary: flow metrics vs previous active month
            if has_activity and prev_active is not None:
                row["mom_new_pct"] = pct_change(new_customers, prev_active["new_customers"])
                row["mom_unique_pct"] = pct_change(
                    unique_m, prev_active["unique_passengers"]
                )
                row["mom_boardings_pct"] = pct_change(
                    boardings_m, prev_active["boardings"]
                )

        prev_year = f"{m0.year - 1}-{m0.month:02d}"
        py = yoy_index.get(prev_year)
        # YoY cumulativo só com base comparável (evita % absurdos em meses quase vazios)
        if (
            py
            and cumulative > 0
            and py.get("cumulative_unique_customers", 0) >= 50
            and (py.get("boardings", 0) >= 50 or py.get("has_activity"))
        ):
            row["yoy_cumulative_pct"] = pct_change(
                cumulative, py["cumulative_unique_customers"]
            )
            row["yoy_vs_month"] = prev_year
            if has_activity and py.get("has_activity") and py.get("boardings", 0) >= 50:
                row["yoy_new_pct"] = pct_change(new_customers, py["new_customers"])
                row["yoy_unique_pct"] = pct_change(unique_m, py["unique_passengers"])
                row["yoy_boardings_pct"] = pct_change(boardings_m, py["boardings"])

        monthly.append(row)
        yoy_index[label] = row
        prev_row = row
        if has_activity:
            prev_active = row

    # Attach LTM deltas vs the same calendar month 12 months earlier
    by_month = {r["month"]: r for r in monthly}
    for r in monthly:
        prev_label = month_minus_years(r["month"], 1)
        prev = by_month.get(prev_label)
        if not prev:
            continue
        r["ltm_vs_month"] = prev_label
        r["ltm_unique_delta_vs_12m"] = _delta(
            r["ltm_unique_customers"], prev["ltm_unique_customers"]
        )
        r["ltm_ge2_delta_vs_12m"] = _delta(r["ltm_ge2"], prev["ltm_ge2"])
        r["ltm_ge4_delta_vs_12m"] = _delta(r["ltm_ge4"], prev["ltm_ge4"])

    active_months = [r for r in monthly if r["has_activity"]]
    last = active_months[-1] if active_months else monthly[-1]
    yoy_points = [r for r in monthly if r["yoy_cumulative_pct"] is not None]

    def _snapshot_from_row(label: str, title: str, row: Optional[dict]) -> dict:
        if not row:
            empty = empty_freq_fields(None)
            return {
                "month": label,
                "label": title,
                "available": False,
                "window_start": None,
                "window_end": None,
                "data_gap": f"sem dados para {label}",
                "ltm_unique_customers": None,
                "ltm_ge2": None,
                "ltm_ge4": None,
                "ltm_ge2_pct": None,
                "ltm_ge4_pct": None,
                **empty,
                "ltm_freq_rows": freq_rows_from_fields(empty),
                "ltm_vs_month": None,
                "ltm_unique_delta_vs_12m": None,
                "ltm_ge2_delta_vs_12m": None,
                "ltm_ge4_delta_vs_12m": None,
                "prev_ltm_ge2": None,
                "prev_ltm_ge4": None,
            }
        prev_label = row.get("ltm_vs_month")
        prev = by_month.get(prev_label) if prev_label else None
        freq = {k: row[k] for k in FREQ_KEYS}
        return {
            "month": label,
            "label": title,
            "available": True,
            "window_start": row["window_start"],
            "window_end": row["window_end"],
            "data_gap": row.get("data_gap"),
            "has_activity": row.get("has_activity"),
            "ltm_unique_customers": row["ltm_unique_customers"],
            "ltm_ge2": row["ltm_ge2"],
            "ltm_ge4": row["ltm_ge4"],
            "ltm_ge2_pct": row["ltm_ge2_pct"],
            "ltm_ge4_pct": row["ltm_ge4_pct"],
            **freq,
            "ltm_freq_rows": freq_rows_from_fields(freq),
            "ltm_vs_month": prev_label,
            "ltm_unique_delta_vs_12m": row.get("ltm_unique_delta_vs_12m"),
            "ltm_ge2_delta_vs_12m": row.get("ltm_ge2_delta_vs_12m"),
            "ltm_ge4_delta_vs_12m": row.get("ltm_ge4_delta_vs_12m"),
            "prev_ltm_ge2": prev["ltm_ge2"] if prev else None,
            "prev_ltm_ge4": prev["ltm_ge4"] if prev else None,
            "prev_ltm_unique": prev["ltm_unique_customers"] if prev else None,
        }

    # Prefer planned Jun/2026 cut when present; else latest active month for headlines
    preferred_latest = by_month.get("2026-06") or last
    snapshots = [
        _snapshot_from_row(label, title, by_month.get(label))
        for label, title in SNAPSHOT_SPECS
    ]

    # Recompute headline KPIs on the filtered window (not raw API all-time)
    uniques_window = len(first_seen)
    repeaters_window = sum(1 for dates in pax_dates.values() if len(dates) >= 2)
    legs_window = sum(len(v) for v in flight_leg_months.values())
    missions_window = sum(mission_month_counts.values())
    boardings_window = len(boardings)
    recurrence_window = (
        round(repeaters_window / uniques_window * 100, 1) if uniques_window else 0.0
    )

    # Frequency over the entire available data period (not rolling LTM)
    period_counts = {pid: len(dates) for pid, dates in pax_dates.items()}
    period_freq = ltm_freq_metrics(period_counts)
    period_frequency = {
        "window_start": max(data_start, base_start).isoformat(),
        "window_end": data_end.isoformat(),
        "unique_customers": period_freq["ltm_unique_customers"],
        "customers_boardings": boardings_window,
        "ge2": period_freq["ltm_ge2"],
        "ge4": period_freq["ltm_ge4"],
        "ge2_pct": period_freq["ltm_ge2_pct"],
        "ge4_pct": period_freq["ltm_ge4_pct"],
        **{k: period_freq[k] for k in FREQ_KEYS},
        "freq_rows": period_freq["ltm_freq_rows"],
    }

    # LTM missions (ending at latest active month) — by mission date
    ltm_mission_total = 0
    if preferred_latest:
        ltm_start = date.fromisoformat(preferred_latest["window_start"])
        ltm_end = date.fromisoformat(preferred_latest["window_end"])
        ltm_mission_total = sum(
            1
            for m in assign_missions(_mission_legs_from_rows(payload_rows, base_start))
            if ltm_start <= m.flight_date <= ltm_end
        )

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "base_start": base_start.isoformat(),
        "data_start": max(data_start, base_start).isoformat(),
        "data_end": data_end.isoformat(),
        "months_available": len(monthly),
        "summary": {
            "unique_customers_all_time": uniques_window,
            "total_boardings": boardings_window,
            "total_flights": missions_window,
            "total_missions": missions_window,
            "total_flight_legs": legs_window,
            "flight_count_unit": "mission",
            "ltm_missions": ltm_mission_total,
            "recurring_all_time": repeaters_window,
            "recurrence_rate_all_time": recurrence_window,
            "period_ge2": period_frequency["ge2"],
            "period_ge4": period_frequency["ge4"],
            "period_ge2_pct": period_frequency["ge2_pct"],
            "period_ge4_pct": period_frequency["ge4_pct"],
            "period_window_start": period_frequency["window_start"],
            "period_window_end": period_frequency["window_end"],
            "cumulative_unique_end": last["cumulative_unique_customers"],
            "latest_month": preferred_latest["month"],
            "latest_mom_cumulative_pct": last["mom_cumulative_pct"],
            "latest_yoy_cumulative_pct": last["yoy_cumulative_pct"],
            "latest_mom_unique_pct": last["mom_unique_pct"],
            "latest_mom_boardings_pct": last["mom_boardings_pct"],
            "latest_yoy_unique_pct": last["yoy_unique_pct"],
            "latest_yoy_boardings_pct": last["yoy_boardings_pct"],
            "ltm_unique_customers": preferred_latest["ltm_unique_customers"],
            "ltm_repeat_customers": preferred_latest["ltm_ge2"],
            "ltm_ge2": preferred_latest["ltm_ge2"],
            "ltm_ge4": preferred_latest["ltm_ge4"],
            "ltm_ge2_pct": preferred_latest["ltm_ge2_pct"],
            "ltm_ge4_pct": preferred_latest["ltm_ge4_pct"],
            "ltm_repeat_rate_pct": preferred_latest["repeat_rate_pct"],
            **{k: preferred_latest[k] for k in FREQ_KEYS},
            "ltm_freq_rows": preferred_latest.get("ltm_freq_rows")
            or freq_rows_from_fields(preferred_latest),
            "ltm_window_start": preferred_latest["window_start"],
            "ltm_window_end": preferred_latest["window_end"],
            "ltm_vs_month": preferred_latest.get("ltm_vs_month"),
            "ltm_unique_delta_vs_12m": preferred_latest.get("ltm_unique_delta_vs_12m"),
            "ltm_ge2_delta_vs_12m": preferred_latest.get("ltm_ge2_delta_vs_12m"),
            "ltm_ge4_delta_vs_12m": preferred_latest.get("ltm_ge4_delta_vs_12m"),
            "prev_ltm_ge2": (
                by_month[preferred_latest["ltm_vs_month"]]["ltm_ge2"]
                if preferred_latest.get("ltm_vs_month") in by_month
                else None
            ),
            "prev_ltm_ge4": (
                by_month[preferred_latest["ltm_vs_month"]]["ltm_ge4"]
                if preferred_latest.get("ltm_vs_month") in by_month
                else None
            ),
            "latest_data_gap": preferred_latest.get("data_gap"),
            "yoy_months_available": len(yoy_points),
            "api_unique_unfiltered": summary.get("unique_passengers"),
        },
        "period_frequency": period_frequency,
        "glossary": [{"term": t, "definition": d} for t, d in GLOSSARY],
        "snapshots": snapshots,
        "monthly": monthly,
        "top_routes": routes,
        "top_passengers": [
            {
                "name": r.get("name"),
                "identity_key": r.get("identity_key", ""),
                "boardings": r.get("boardings", 0),
                "distinct_dates": r.get("distinct_dates", 0),
                "first_in_window": r.get("first_in_window") or "",
                "last_in_window": r.get("last_in_window") or "",
            }
            for r in top
        ],
    }


HTML = r'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>REVO · Recorrência LTM</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
  <script src="./data.js"></script>
  <style>
    :root {
      --bg: #eef2ec; --ink: #142018; --muted: #5a675c; --panel: #fffef9;
      --line: #d5ddd6; --a: #0b6b52; --b: #c45c26; --c: #2f5d9f;
      --font: "Iowan Old Style", Palatino, Georgia, serif;
      --sans: "Avenir Next", "Segoe UI", sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0; color: var(--ink); font-family: var(--sans);
      background:
        radial-gradient(ellipse 60% 40% at 100% 0%, #dceadf, transparent 55%),
        radial-gradient(ellipse 50% 35% at 0% 100%, #efe4d6, transparent 50%),
        var(--bg);
    }
    main { max-width: 1180px; margin: 0 auto; padding: 36px 20px 80px; }
    h1 { font-family: var(--font); font-size: clamp(2.2rem, 5vw, 3.3rem); margin: 0; letter-spacing: -0.02em; }
    .lede { color: var(--muted); max-width: 46em; line-height: 1.5; margin: 10px 0 0; }
    .meta { margin-top: 12px; color: var(--muted); font-size: 0.9rem; }
    .meta a { color: var(--a); }
    .alert {
      margin: 16px 0 0; padding: 10px 12px; border: 1px solid #e2c49a;
      background: #fff6e8; color: #6a4a1a; font-size: 0.9rem;
    }
    .kpis { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin: 26px 0; }
    .kpis article { background: var(--panel); border: 1px solid var(--line); padding: 14px 16px; }
    .kpis .label { display: block; font-size: 0.7rem; letter-spacing: 0.08em; text-transform: uppercase; color: var(--muted); margin-bottom: 6px; }
    .kpis strong { font-family: var(--font); font-size: 1.7rem; }
    .kpis .sub { display: block; margin-top: 4px; color: var(--muted); font-size: 0.85rem; }
    section { margin-top: 28px; }
    h2 { font-family: var(--font); font-size: 1.35rem; margin: 0 0 6px; }
    .help { color: var(--muted); margin: 0 0 14px; font-size: 0.95rem; }
    .charts { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
    .charts.full { grid-template-columns: 1fr; }
    .box { background: var(--panel); border: 1px solid var(--line); padding: 14px; }
    .box h3 { margin: 0 0 10px; color: var(--muted); font-size: 0.95rem; }
    .table-wrap { overflow-x: auto; background: var(--panel); border: 1px solid var(--line); }
    table { width: 100%; border-collapse: collapse; font-size: 0.86rem; }
    th, td { padding: 9px 10px; border-bottom: 1px solid var(--line); text-align: left; white-space: nowrap; }
    th { color: var(--muted); font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.05em; }
    td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
    .pos { color: #0b6b52; } .neg { color: #b33b2b; } .na { color: #9aa39b; }
    @media (max-width: 900px) { .kpis, .charts { grid-template-columns: 1fr; } }
  </style>
</head>
<body>
<main>
  <header>
    <h1>REVO</h1>
    <p class="lede">Recorrência LTM (últimos 12 meses): quantos passageiros voaram 1× … 20× e &gt;20. Cortes-chave ≥2 e ≥4, com delta vs o LTM de 12 meses atrás. SIAV→SIAV excluídos.</p>
    <p class="meta" id="meta"></p>
    <p class="meta"><a href="./revo-customer-kpis.xlsx">Baixar Excel</a></p>
    <p class="alert" id="gapAlert" hidden></p>
  </header>

  <section>
    <h2>Glossário</h2>
    <p class="help">Definições das métricas usadas neste dashboard.</p>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Termo</th>
            <th>Significado</th>
          </tr>
        </thead>
        <tbody id="glossaryBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Base total</h2>
    <p class="help">Unique customers = passageiros distintos. Customers = boardings (cada embarque conta).</p>
    <div class="kpis" id="baseKpis"></div>
  </section>

  <section id="sfRevenueSection" hidden>
    <h2>Base de vendas · corporate mobility</h2>
    <p class="help" id="sfRevenueHelp">Evolução do faturamento (mês do voo) e quanto disso é corporate mobility (PJ).</p>
    <div class="kpis" id="sfRevenueKpis"></div>
    <div class="charts" style="margin:14px 0">
      <div class="box"><h3>Evolução da base de vendas</h3><canvas id="chartSales" height="240"></canvas></div>
      <div class="box"><h3>Corporate mobility % do faturamento</h3><canvas id="chartCorp" height="240"></canvas></div>
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Mês</th>
            <th class="num">Faturamento</th>
            <th class="num">Corporate PJ</th>
            <th class="num">Corporate %</th>
          </tr>
        </thead>
        <tbody id="sfRevMonthBody"></tbody>
      </table>
    </div>
  </section>

  <section id="sfHoursSection" hidden>
    <h2>Shuttle equivalent · horas voadas (Salesforce)</h2>
    <p class="help" id="sfHoursHelp"></p>
    <div class="kpis" id="sfHoursKpis"></div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Mês</th>
            <th class="num">Horas total</th>
            <th class="num">SBGR</th>
            <th class="num">Resto</th>
            <th class="num">% SBGR</th>
            <th class="num">Shuttle</th>
            <th class="num">Shuttle∩SBGR</th>
            <th class="num">Voos</th>
          </tr>
        </thead>
        <tbody id="sfHoursBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Recorrência LTM</h2>
    <p class="help" id="ltmHelp"></p>
    <div class="kpis" id="recKpis"></div>
  </section>

  <section>
    <h2>Frequência LTM · mês a mês</h2>
    <p class="help">Para cada mês-fim, distribuição do LTM (até 12 meses) em 1× … 20× e &gt;20, com unique LTM e customers (boardings) do mês civil.</p>
    <div class="table-wrap">
      <table>
        <thead id="freqMonthHead"></thead>
        <tbody id="freqMonthBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Frequência · todo o período</h2>
    <p class="help" id="periodFreqHelp">Distribuição 1× … 20× e &gt;20 em toda a base disponível (não é LTM).</p>
    <div class="kpis" id="periodKpis"></div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Frequência</th>
            <th class="num">Passageiros</th>
            <th class="num">% dos unique</th>
          </tr>
        </thead>
        <tbody id="periodFreqBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Distribuição por frequência (LTM atual)</h2>
    <p class="help">Passageiros únicos no LTM agrupados pelo nº de boardings na janela (1× a 20× e &gt;20).</p>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Frequência</th>
            <th class="num">Passageiros</th>
            <th class="num">% do LTM</th>
          </tr>
        </thead>
        <tbody id="freqBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Snapshots · Jun/2026 · Dez/2025 · Dez/2024</h2>
    <p class="help">Mesma métrica LTM em três cortes de referência. Delta vs o LTM que terminava 12 meses antes.</p>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Snapshot</th>
            <th>Janela LTM</th>
            <th class="num">Unique</th>
            <th class="num">≥2</th>
            <th class="num">Δ ≥2</th>
            <th class="num">≥4</th>
            <th class="num">Δ ≥4</th>
          </tr>
        </thead>
        <tbody id="snapBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Frequência por snapshot (1× … 20× e &gt;20)</h2>
    <p class="help">Comparação da distribuição completa nos três cortes.</p>
    <div class="table-wrap">
      <table>
        <thead id="freqSnapHead"></thead>
        <tbody id="freqSnapBody"></tbody>
      </table>
    </div>
  </section>

  <div class="kpis" id="kpis"></div>

  <section>
    <h2>Recorrentes ≥2 e ≥4 (rolling LTM)</h2>
    <div class="charts">
      <div class="box"><h3>Passageiros ≥2 / ≥4 no LTM</h3><canvas id="chartGe" height="240"></canvas></div>
      <div class="box"><h3>Taxa ≥2 % (repeat rate)</h3><canvas id="chartRepeat" height="240"></canvas></div>
    </div>
  </section>

  <section>
    <h2>Cumulativo unique customers</h2>
    <p class="help">Base acumulada de clientes únicos desde janeiro/2024 (sem SIAV→SIAV).</p>
    <div class="charts full"><div class="box"><canvas id="chartCum" height="260"></canvas></div></div>
  </section>

  <section>
    <h2>Variação do cumulativo · MoM</h2>
    <p class="help">Crescimento percentual do cumulativo de unique vs o mês anterior.</p>
    <div class="charts full">
      <div class="box"><h3>MoM % · cumulativo unique</h3><canvas id="chartMomCum" height="260"></canvas></div>
    </div>
  </section>

  <section>
    <h2>Variação do cumulativo · YoY</h2>
    <p class="help">Crescimento percentual do cumulativo vs o mesmo mês do ano anterior (quando existir).</p>
    <div class="charts full">
      <div class="box"><h3>YoY % · cumulativo unique</h3><canvas id="chartYoyCum" height="260"></canvas></div>
    </div>
  </section>

  <section>
    <h2>Novos clientes / mês</h2>
    <div class="charts full">
      <div class="box"><canvas id="chartNew" height="240"></canvas></div>
    </div>
  </section>

  <section>
    <h2>Qualidade da base</h2>
    <p class="help" id="quality"></p>
  </section>

  <section>
    <h2>Série mensal completa</h2>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Mês</th>
            <th class="num">Novos</th>
            <th class="num">Cumulativo</th>
            <th class="num">Unique mês</th>
            <th class="num">Customers</th>
            <th class="num">LTM unique</th>
            <th class="num">LTM ≥2</th>
            <th class="num">Δ ≥2</th>
            <th class="num">LTM ≥4</th>
            <th class="num">Δ ≥4</th>
            <th class="num">≥2 %</th>
            <th class="num">MoM cumul.%</th>
            <th class="num">YoY cumul.%</th>
          </tr>
        </thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </section>
</main>
<script>
const D = window.KPI_DATA || {};
const s = D.summary || {};
const monthly = D.monthly || [];
const snapshots = D.snapshots || [];
const periodFreq = D.period_frequency || {};
const glossary = D.glossary || [];
const sf = D.salesforce || null;
const fmt = (v, suffix='%') => v == null ? '—' : `${v > 0 ? '+' : ''}${v}${suffix}`;
const fmtN = (v) => v == null ? '—' : String(v);
const fmtDelta = (v) => v == null ? '—' : `${v > 0 ? '+' : ''}${v}`;
const cls = (v) => v == null ? 'na' : (v >= 0 ? 'pos' : 'neg');
const pctOf = (n, den) => (!den || n == null) ? '—' : `${(n / den * 100).toFixed(1)}%`;
const fmtBRL = (v) => v == null ? '—' : Number(v).toLocaleString('pt-BR', {style:'currency', currency:'BRL', maximumFractionDigits:0});
const fmtH = (v) => v == null ? '—' : Number(v).toLocaleString('pt-BR', {maximumFractionDigits:1});

document.getElementById('glossaryBody').innerHTML = glossary.map(g => `
  <tr>
    <td><strong>${g.term}</strong></td>
    <td style="white-space:normal;max-width:52em">${g.definition}</td>
  </tr>`).join('');

if (sf && sf.hours) {
  document.getElementById('sfHoursSection').hidden = false;
  const h = sf.hours;
  const max = h.historic_max || {};
  document.getElementById('sfHoursHelp').textContent =
    `${h.method || ''} · Máx histórico: total ${max.all?.month} ${fmtH(max.all?.hours)}h · SBGR ${max.sbgr?.month} ${fmtH(max.sbgr?.hours)}h · Shuttle ${max.shuttle?.month} ${fmtH(max.shuttle?.hours)}h`;
  const latest = h.latest || {};
  document.getElementById('sfHoursKpis').innerHTML = [
    ['Horas (jun/2026)', fmtH(latest.hours_all), `${fmtN(latest.flights)} voos executados`],
    ['SBGR', fmtH(latest.hours_sbgr), `${pctOf(latest.hours_sbgr, latest.hours_all)} do total`],
    ['Resto', fmtH(latest.hours_resto), `Shuttle ${fmtH(latest.hours_shuttle)}h · ∩SBGR ${fmtH(latest.hours_shuttle_sbgr)}h`],
  ].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${v}</strong><span class="sub">${sub||''}</span></article>`).join('');
  const months = Object.keys(h.monthly || {}).sort();
  document.getElementById('sfHoursBody').innerHTML = months.map(m => {
    const r = h.monthly[m];
    const mark = (max.all?.month===m || max.sbgr?.month===m || max.shuttle?.month===m) ? ' ★' : '';
    return `<tr>
      <td>${m}${mark}</td>
      <td class="num">${fmtH(r.hours_all)}</td>
      <td class="num">${fmtH(r.hours_sbgr)}</td>
      <td class="num">${fmtH(r.hours_resto)}</td>
      <td class="num">${pctOf(r.hours_sbgr, r.hours_all)}</td>
      <td class="num">${fmtH(r.hours_shuttle)}</td>
      <td class="num">${fmtH(r.hours_shuttle_sbgr)}</td>
      <td class="num">${fmtN(r.flights)}</td>
    </tr>`;
  }).join('');
}

if (sf && sf.revenue) {
  document.getElementById('sfRevenueSection').hidden = false;
  const rev = sf.revenue;
  const snap = rev.snapshots?.ltm_2026_06 || {};
  const corpPct = snap.corporate_pj_pago_pct ?? snap.corporate_pj_pct;
  const corpAmt = snap.corporate_pj_pago ?? snap.corporate_pj;
  document.getElementById('sfRevenueKpis').innerHTML = [
    ['Faturamento LTM Jun/2026', fmtBRL(snap.valor_pago), 'base de vendas · mês do voo'],
    ['Corporate mobility', `${corpPct ?? '—'}%`, fmtBRL(corpAmt) + ' do faturamento LTM'],
  ].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${v}</strong><span class="sub">${sub||''}</span></article>`).join('');
  const rm = Object.keys(rev.monthly || {}).sort();
  const salesRows = rm.map(m => {
    const r = rev.monthly[m];
    const fat = r.valor_pago;
    const corp = r.corporate_pj_pago ?? r.corporate_pj;
    const pct = r.corporate_pj_pago_pct ?? r.corporate_pj_pct;
    return { m, fat, corp, pct };
  });
  document.getElementById('sfRevMonthBody').innerHTML = salesRows.map(r => `
    <tr>
      <td>${r.m}</td>
      <td class="num">${fmtBRL(r.fat)}</td>
      <td class="num">${fmtBRL(r.corp)}</td>
      <td class="num">${r.pct ?? '—'}%</td>
    </tr>`).join('');
  window.__salesRows = salesRows;
}

const gaps = monthly.filter(r => r.data_gap);
document.getElementById('quality').innerHTML = gaps.length
  ? ('Lacunas de ingestão: ' + gaps.map(r => `<strong>${r.month}</strong> — ${r.data_gap}`).join(' · ') + '. Suba os Excel faltantes na API de Manifests para preencher.')
  : 'Nenhuma lacuna conhecida no período.';

if (s.latest_data_gap) {
  const el = document.getElementById('gapAlert');
  el.hidden = false;
  el.textContent = `Atenção no LTM atual (${s.latest_month}): ${s.latest_data_gap}`;
}

document.getElementById('meta').textContent =
  `Fonte: ${D.source || '—'} · base ${D.base_start || D.data_start || '—'} → ${D.data_end || '—'} · ${D.months_available || 0} meses · gerado ${D.generated_at || '—'}`;

const winLabel = (s.ltm_window_start && s.ltm_window_end)
  ? `${s.ltm_window_start} → ${s.ltm_window_end}`
  : '—';
document.getElementById('ltmHelp').textContent =
  `Janela LTM atual (${s.latest_month || '—'}): ${winLabel}. Comparado ao LTM que terminava em ${s.ltm_vs_month || '—'}.`;

document.getElementById('baseKpis').innerHTML = [
  ['Unique customers', s.unique_customers_all_time, `base ${D.base_start || 'jan/2024'} → ${D.data_end || '—'} · cumulativo ${fmtN(s.cumulative_unique_end)}`],
  ['Customers (boardings)', s.total_boardings, `cada embarque conta · missões ${fmtN(s.total_missions ?? s.total_flights)}`],
  ['Unique LTM', s.ltm_unique_customers, `janela ${winLabel}`],
].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${fmtN(v)}</strong><span class="sub">${sub || ''}</span></article>`).join('');

document.getElementById('recKpis').innerHTML = [
  ['Unique LTM', s.ltm_unique_customers, `Δ ${fmtDelta(s.ltm_unique_delta_vs_12m)} vs ${s.ltm_vs_month || '—'}`],
  ['Recorrentes ≥2', s.ltm_ge2, `há 12m: ${fmtN(s.prev_ltm_ge2)} · hoje ${fmtDelta(s.ltm_ge2_delta_vs_12m)} · ${s.ltm_ge2_pct ?? '—'}%`],
  ['Recorrentes ≥4', s.ltm_ge4, `há 12m: ${fmtN(s.prev_ltm_ge4)} · hoje ${fmtDelta(s.ltm_ge4_delta_vs_12m)} · ${s.ltm_ge4_pct ?? '—'}%`],
].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${fmtN(v)}</strong><span class="sub">${sub || ''}</span></article>`).join('');

const periodRows = periodFreq.freq_rows || [];
document.getElementById('periodFreqHelp').textContent =
  `Distribuição 1× … 20× e >20 em toda a base (${periodFreq.window_start || D.data_start || '—'} → ${periodFreq.window_end || D.data_end || '—'}). Não é LTM.`;
document.getElementById('periodKpis').innerHTML = [
  ['Unique (período)', periodFreq.unique_customers, `customers ${fmtN(periodFreq.customers_boardings)}`],
  ['≥2 no período', periodFreq.ge2, `${periodFreq.ge2_pct ?? '—'}%`],
  ['≥4 no período', periodFreq.ge4, `${periodFreq.ge4_pct ?? '—'}%`],
].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${fmtN(v)}</strong><span class="sub">${sub || ''}</span></article>`).join('');
document.getElementById('periodFreqBody').innerHTML = periodRows.map(r => `
  <tr>
    <td>${r.label}</td>
    <td class="num">${fmtN(r.count)}</td>
    <td class="num">${pctOf(r.count, periodFreq.unique_customers)}</td>
  </tr>`).join('');

const freqRows = s.ltm_freq_rows || [];
const freqKeys = freqRows.map(r => r.key);
document.getElementById('freqMonthHead').innerHTML = `<tr>
  <th>Mês</th>
  <th class="num">Unique mês</th>
  <th class="num">Customers</th>
  <th class="num">Unique LTM</th>
  <th class="num">≥2</th>
  <th class="num">≥4</th>
  ${freqRows.map(r => `<th class="num">${r.label}</th>`).join('')}
</tr>`;
document.getElementById('freqMonthBody').innerHTML = monthly.map(r => `
  <tr>
    <td>${r.month}${r.data_gap ? ' ⚠' : (r.has_activity ? '' : ' <span class="na">(vazio)</span>')}</td>
    <td class="num">${fmtN(r.unique_passengers)}</td>
    <td class="num">${fmtN(r.boardings)}</td>
    <td class="num">${fmtN(r.ltm_unique_customers)}</td>
    <td class="num">${fmtN(r.ltm_ge2)}</td>
    <td class="num">${fmtN(r.ltm_ge4)}</td>
    ${freqKeys.map(k => `<td class="num">${fmtN(r[k])}</td>`).join('')}
  </tr>`).join('');

document.getElementById('freqBody').innerHTML = freqRows.map(r => `
  <tr>
    <td>${r.label}</td>
    <td class="num">${fmtN(r.count)}</td>
    <td class="num">${pctOf(r.count, s.ltm_unique_customers)}</td>
  </tr>`).join('');

document.getElementById('snapBody').innerHTML = snapshots.map(r => {
  if (!r.available) {
    return `<tr><td>${r.label}</td><td colspan="6" class="na">${r.data_gap || 'sem dados'}</td></tr>`;
  }
  const win = (r.window_start && r.window_end) ? `${r.window_start} → ${r.window_end}` : '—';
  const gap = r.data_gap ? ` <span class="na">⚠</span>` : '';
  return `<tr>
    <td>${r.label}${gap}</td>
    <td>${win}</td>
    <td class="num">${fmtN(r.ltm_unique_customers)}</td>
    <td class="num">${fmtN(r.ltm_ge2)}</td>
    <td class="num ${cls(r.ltm_ge2_delta_vs_12m)}">${fmtDelta(r.ltm_ge2_delta_vs_12m)}</td>
    <td class="num">${fmtN(r.ltm_ge4)}</td>
    <td class="num ${cls(r.ltm_ge4_delta_vs_12m)}">${fmtDelta(r.ltm_ge4_delta_vs_12m)}</td>
  </tr>`;
}).join('');

document.getElementById('freqSnapHead').innerHTML = `<tr>
  <th>Frequência</th>
  ${snapshots.map(r => `<th class="num">${r.label}</th>`).join('')}
</tr>`;
const freqTemplate = (snapshots.find(r => (r.ltm_freq_rows || []).length) || {}).ltm_freq_rows || freqRows;
document.getElementById('freqSnapBody').innerHTML = freqTemplate.map((row, idx) => `
  <tr>
    <td>${row.label}</td>
    ${snapshots.map(sn => {
      const n = (sn.ltm_freq_rows && sn.ltm_freq_rows[idx]) ? sn.ltm_freq_rows[idx].count : sn[row.key];
      return `<td class="num">${fmtN(n)}</td>`;
    }).join('')}
  </tr>`).join('');

document.getElementById('kpis').innerHTML = [
  ['Unique customers', s.unique_customers_all_time, `customers ${fmtN(s.total_boardings)}`],
  ['Missões (corte Sigtrip)', s.total_missions ?? s.total_flights, `${s.total_flight_legs || '—'} pernas · LTM ${s.ltm_missions ?? '—'}`],
  ['Cumulativo final', s.cumulative_unique_end, `até ${s.latest_month || '—'}`],
  ['MoM cumulativo', fmt(s.latest_mom_cumulative_pct), `último mês ${s.latest_month || '—'}`],
  ['YoY cumulativo', fmt(s.latest_yoy_cumulative_pct), s.yoy_months_available ? `${s.yoy_months_available} meses com YoY` : 'sem par YoY ainda'],
  ['LTM ≥2 / ≥4', `${fmtN(s.ltm_ge2)} / ${fmtN(s.ltm_ge4)}`, `unique LTM ${fmtN(s.ltm_unique_customers)}`],
].map(([l,v,sub]) => `<article><span class="label">${l}</span><strong>${v ?? '—'}</strong><span class="sub">${sub || ''}</span></article>`).join('');

document.getElementById('tbody').innerHTML = monthly.map(r => `
  <tr>
    <td>${r.month}${r.data_gap ? ` <span class="na">⚠ ${r.data_gap}</span>` : (r.has_activity ? '' : ' <span class="na">(sem voos)</span>')}</td>
    <td class="num">${r.new_customers}</td>
    <td class="num">${r.cumulative_unique_customers}</td>
    <td class="num">${r.unique_passengers}</td>
    <td class="num">${r.boardings}</td>
    <td class="num">${r.ltm_unique_customers}</td>
    <td class="num">${r.ltm_ge2}</td>
    <td class="num ${cls(r.ltm_ge2_delta_vs_12m)}">${fmtDelta(r.ltm_ge2_delta_vs_12m)}</td>
    <td class="num">${r.ltm_ge4}</td>
    <td class="num ${cls(r.ltm_ge4_delta_vs_12m)}">${fmtDelta(r.ltm_ge4_delta_vs_12m)}</td>
    <td class="num">${r.ltm_ge2_pct}%</td>
    <td class="num ${cls(r.mom_cumulative_pct)}">${fmt(r.mom_cumulative_pct)}</td>
    <td class="num ${cls(r.yoy_cumulative_pct)}">${fmt(r.yoy_cumulative_pct)}</td>
  </tr>`).join('');

const labels = monthly.map(r => r.month);
const active = monthly.filter(r => r.has_activity);
const momCum = monthly.filter(r => r.mom_cumulative_pct != null);
const yoyCum = monthly.filter(r => r.yoy_cumulative_pct != null);
const grid = { color: 'rgba(20,32,24,0.08)' };
const line = (id, datasets, opts={}) => new Chart(document.getElementById(id), {
  type: 'line',
  data: { labels: opts.labels || labels, datasets },
  options: {
    plugins: { legend: { display: opts.legend !== false, position: 'bottom' } },
    scales: {
      x: { grid },
      y: {
        grid,
        beginAtZero: !!opts.beginAtZero,
        ticks: opts.pct ? { callback: v => v + '%' } : undefined,
      },
    },
  },
});

line('chartGe', [
  {
    label: '≥2',
    data: active.map(r => r.ltm_ge2),
    borderColor: '#0b6b52', backgroundColor: 'rgba(11,107,82,0.10)', fill: false, tension: 0.25, pointRadius: 3,
  },
  {
    label: '≥4',
    data: active.map(r => r.ltm_ge4),
    borderColor: '#c45c26', backgroundColor: 'rgba(196,92,38,0.10)', fill: false, tension: 0.25, pointRadius: 3,
  },
], { labels: active.map(r => r.month), beginAtZero: true, legend: true });

line('chartRepeat', [{
  label: '≥2 %',
  data: active.map(r => r.ltm_ge2_pct ?? r.repeat_rate_pct),
  borderColor: '#c45c26', backgroundColor: 'rgba(196,92,38,0.10)', fill: true, tension: 0.25, pointRadius: 3,
}], { labels: active.map(r => r.month), pct: true, beginAtZero: true, legend: false });

line('chartCum', [{
  label: 'Cumulativo unique',
  data: monthly.map(r => r.cumulative_unique_customers),
  borderColor: '#0b6b52', backgroundColor: 'rgba(11,107,82,0.12)', fill: true, tension: 0.25, pointRadius: 3,
}], { beginAtZero: true, legend: false });

line('chartMomCum', [{
  label: 'MoM cumulativo %',
  data: momCum.map(r => r.mom_cumulative_pct),
  borderColor: '#c45c26', backgroundColor: 'rgba(196,92,38,0.12)', fill: true, tension: 0.2, pointRadius: 3, spanGaps: true,
}], { labels: momCum.map(r => r.month), pct: true, legend: false });

line('chartYoyCum', [{
  label: 'YoY cumulativo %',
  data: yoyCum.map(r => r.yoy_cumulative_pct),
  borderColor: '#2f5d9f', backgroundColor: 'rgba(47,93,159,0.12)', fill: true, tension: 0.2, pointRadius: 3, spanGaps: true,
}], { labels: yoyCum.map(r => r.month), pct: true, legend: false });

line('chartNew', [{
  label: 'Novos',
  data: active.map(r => r.new_customers),
  borderColor: '#0b6b52', tension: 0.25, pointRadius: 3,
}], { labels: active.map(r => r.month), beginAtZero: true, legend: false });

if (window.__salesRows && window.__salesRows.length) {
  const sr = window.__salesRows;
  line('chartSales', [{
    label: 'Faturamento',
    data: sr.map(r => r.fat),
    borderColor: '#0b6b52', backgroundColor: 'rgba(11,107,82,0.12)', fill: true, tension: 0.25, pointRadius: 3,
  }], { labels: sr.map(r => r.m), beginAtZero: true, legend: false });
  line('chartCorp', [{
    label: 'Corporate %',
    data: sr.map(r => r.pct),
    borderColor: '#c45c26', backgroundColor: 'rgba(196,92,38,0.12)', fill: true, tension: 0.25, pointRadius: 3,
  }], { labels: sr.map(r => r.m), pct: true, beginAtZero: true, legend: false });
}
</script>
</body>
</html>
'''


def load_salesforce_kpis() -> Optional[dict]:
    path = OUT / "salesforce_kpis.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"Could not load salesforce_kpis.json: {exc}")
        return None


def write_excel(data: dict, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font

    wb = Workbook()
    s = data["summary"]
    sf = data.get("salesforce") or {}

    ws_g = wb.active
    ws_g.title = "Glossário"
    ws_g["A1"] = "REVO · Glossário de métricas"
    ws_g["A1"].font = Font(size=16, bold=True)
    ws_g["A2"] = "Definições usadas no dashboard de recorrência / customer KPIs"
    ws_g.cell(4, 1, "Termo")
    ws_g.cell(4, 2, "Significado")
    ws_g.cell(4, 1).font = Font(bold=True)
    ws_g.cell(4, 2).font = Font(bold=True)
    for r_i, (term, definition) in enumerate(GLOSSARY, 5):
        ws_g.cell(r_i, 1, term)
        ws_g.cell(r_i, 2, definition)
    ws_g.column_dimensions["A"].width = 28
    ws_g.column_dimensions["B"].width = 92

    if sf.get("hours"):
        ws_h = wb.create_sheet("Horas SBGR")
        ws_h["A1"] = "Shuttle equivalent · horas voadas (Salesforce)"
        ws_h["A1"].font = Font(size=14, bold=True)
        ws_h["A2"] = sf["hours"].get("method") or ""
        mx = sf["hours"].get("historic_max") or {}
        ws_h["A3"] = (
            f"Máx total {mx.get('all',{}).get('month')} {mx.get('all',{}).get('hours')}h · "
            f"SBGR {mx.get('sbgr',{}).get('month')} {mx.get('sbgr',{}).get('hours')}h · "
            f"Shuttle {mx.get('shuttle',{}).get('month')} {mx.get('shuttle',{}).get('hours')}h"
        )
        headers_h = [
            "Mês", "Horas total", "SBGR", "Resto", "% SBGR",
            "Shuttle", "Shuttle∩SBGR", "Voos", "Voos SBGR", "Voos Shuttle",
        ]
        for i, h in enumerate(headers_h, 1):
            ws_h.cell(5, i, h)
        for r_i, m in enumerate(sorted((sf["hours"].get("monthly") or {})), 6):
            r = sf["hours"]["monthly"][m]
            tot = r.get("hours_all") or 0
            vals = [
                m, r.get("hours_all"), r.get("hours_sbgr"), r.get("hours_resto"),
                round(r.get("hours_sbgr", 0) / tot * 100, 1) if tot else 0,
                r.get("hours_shuttle"), r.get("hours_shuttle_sbgr"),
                r.get("flights"), r.get("flights_sbgr"), r.get("flights_shuttle"),
            ]
            for c, v in enumerate(vals, 1):
                ws_h.cell(r_i, c, v)

    if sf.get("revenue"):
        ws_rv = wb.create_sheet("Base de vendas")
        ws_rv["A1"] = "Base de vendas · corporate mobility"
        ws_rv["A1"].font = Font(size=14, bold=True)
        ws_rv["A2"] = (
            "Faturamento = Servico.ValorPago no mês do voo · "
            "Corporate % = Conta Faturamento Pessoa Jurídica / faturamento"
        )
        snap = (sf["revenue"].get("snapshots") or {}).get("ltm_2026_06") or {}
        ws_rv["A4"] = "LTM Jun/2026"
        ws_rv["A4"].font = Font(bold=True)
        ws_rv["A5"] = "Faturamento"
        ws_rv["B5"] = snap.get("valor_pago")
        ws_rv["A6"] = "Corporate mobility %"
        ws_rv["B6"] = snap.get("corporate_pj_pago_pct")
        ws_rv["C6"] = snap.get("corporate_pj_pago")
        ws_rv["A8"] = "Mensal"
        ws_rv["A8"].font = Font(bold=True)
        for i, h in enumerate(
            ["Mês", "Faturamento", "Corporate PJ", "Corporate %"],
            1,
        ):
            ws_rv.cell(9, i, h)
        for r_i, m in enumerate(sorted((sf["revenue"].get("monthly") or {})), 10):
            r = sf["revenue"]["monthly"][m]
            vals = [
                m,
                r.get("valor_pago"),
                r.get("corporate_pj_pago"),
                r.get("corporate_pj_pago_pct"),
            ]
            for c, v in enumerate(vals, 1):
                ws_rv.cell(r_i, c, v if v is not None else None)

    ws_rec = wb.create_sheet("Recorrência LTM")
    ws_rec["A1"] = "REVO · Recorrência LTM"
    ws_rec["A1"].font = Font(size=16, bold=True)
    ws_rec["A2"] = (
        f"Fonte {data['source']} · base {data.get('base_start') or data['data_start']} "
        f"→ {data['data_end']} · LTM {s.get('ltm_window_start')} → {s.get('ltm_window_end')}"
    )
    if s.get("latest_data_gap"):
        ws_rec["A3"] = f"Alerta: {s['latest_month']} — {s['latest_data_gap']}"

    headline = [
        ("Unique customers", s.get("unique_customers_all_time")),
        ("Customers (boardings)", s.get("total_boardings")),
        ("Unique LTM", s.get("ltm_unique_customers")),
        ("Δ unique vs −12m", s.get("ltm_unique_delta_vs_12m")),
        ("≥2 LTM", s.get("ltm_ge2")),
        ("≥2 há 12m", s.get("prev_ltm_ge2")),
        ("Δ ≥2", s.get("ltm_ge2_delta_vs_12m")),
        ("≥2 %", s.get("ltm_ge2_pct")),
        ("≥4 LTM", s.get("ltm_ge4")),
        ("≥4 há 12m", s.get("prev_ltm_ge4")),
        ("Δ ≥4", s.get("ltm_ge4_delta_vs_12m")),
        ("≥4 %", s.get("ltm_ge4_pct")),
    ]
    for i, (lab, val) in enumerate(headline):
        ws_rec.cell(5, 1 + i, lab)
        ws_rec.cell(6, 1 + i, val if val is not None else "—")

    period = data.get("period_frequency") or {}
    period_rows = period.get("freq_rows") or []
    ws_rec["A8"] = (
        f"Frequência · todo o período "
        f"({period.get('window_start') or data.get('data_start')} → "
        f"{period.get('window_end') or data.get('data_end')})"
    )
    ws_rec["A8"].font = Font(bold=True)
    ws_rec["A9"] = (
        f"Unique {period.get('unique_customers')} · Customers {period.get('customers_boardings')} · "
        f"≥2 {period.get('ge2')} ({period.get('ge2_pct')}%) · ≥4 {period.get('ge4')} ({period.get('ge4_pct')}%)"
    )
    for i, h in enumerate(["Frequência", "Passageiros", "% dos unique"], 1):
        ws_rec.cell(10, i, h)
    unique_period = period.get("unique_customers") or 0
    for r_i, row in enumerate(period_rows, 11):
        n = row.get("count")
        ws_rec.cell(r_i, 1, row.get("label"))
        ws_rec.cell(r_i, 2, n if n is not None else "—")
        if n is not None and unique_period:
            ws_rec.cell(r_i, 3, round(n / unique_period * 100, 1))
        else:
            ws_rec.cell(r_i, 3, "—")

    ltm_title_row = 11 + len(period_rows) + 1
    ws_rec.cell(ltm_title_row, 1, "Distribuição por frequência (LTM atual · 1× … 20× e >20)")
    ws_rec.cell(ltm_title_row, 1).font = Font(bold=True)
    for i, h in enumerate(["Frequência", "Passageiros", "% do LTM"], 1):
        ws_rec.cell(ltm_title_row + 1, i, h)
    unique_ltm = s.get("ltm_unique_customers") or 0
    freq_rows = s.get("ltm_freq_rows") or freq_rows_from_fields(s)
    ltm_data_start = ltm_title_row + 2
    for r_i, row in enumerate(freq_rows, ltm_data_start):
        n = row.get("count")
        ws_rec.cell(r_i, 1, row.get("label"))
        ws_rec.cell(r_i, 2, n if n is not None else "—")
        if n is not None and unique_ltm:
            ws_rec.cell(r_i, 3, round(n / unique_ltm * 100, 1))
        else:
            ws_rec.cell(r_i, 3, "—")

    snap_start = ltm_data_start + len(freq_rows) + 2
    ws_rec.cell(snap_start, 1, "Snapshots (Jun/2026 · Dez/2025 · Dez/2024)")
    ws_rec.cell(snap_start, 1).font = Font(bold=True)
    snap_headers = [
        "Snapshot", "Mês", "Janela início", "Janela fim", "Unique",
        "≥2", "≥2 há 12m", "Δ ≥2", "≥4", "≥4 há 12m", "Δ ≥4", "Alerta",
    ] + FREQ_LABELS
    header_row = snap_start + 1
    for i, h in enumerate(snap_headers, 1):
        ws_rec.cell(header_row, i, h)
    for r_i, snap in enumerate(data.get("snapshots") or [], header_row + 1):
        vals = [
            snap.get("label"),
            snap.get("month"),
            snap.get("window_start"),
            snap.get("window_end"),
            snap.get("ltm_unique_customers"),
            snap.get("ltm_ge2"),
            snap.get("prev_ltm_ge2"),
            snap.get("ltm_ge2_delta_vs_12m"),
            snap.get("ltm_ge4"),
            snap.get("prev_ltm_ge4"),
            snap.get("ltm_ge4_delta_vs_12m"),
            snap.get("data_gap") or "",
        ] + [snap.get(k) for k in FREQ_KEYS]
        for c, v in enumerate(vals, 1):
            ws_rec.cell(r_i, c, v if v is not None else "—")

    # Cross-tab: frequency rows × snapshot columns
    xtab_start = header_row + len(data.get("snapshots") or []) + 3
    ws_rec.cell(xtab_start, 1, "Frequência por snapshot")
    ws_rec.cell(xtab_start, 1).font = Font(bold=True)
    snaps = data.get("snapshots") or []
    ws_rec.cell(xtab_start + 1, 1, "Frequência")
    for c, snap in enumerate(snaps, 2):
        ws_rec.cell(xtab_start + 1, c, snap.get("label"))
    for r_i, row in enumerate(freq_rows, xtab_start + 2):
        ws_rec.cell(r_i, 1, row.get("label"))
        for c, snap in enumerate(snaps, 2):
            n = snap.get(row.get("key"))
            ws_rec.cell(r_i, c, n if n is not None else "—")

    ws_fm = wb.create_sheet("Freq LTM Mensal")
    ws_fm["A1"] = "Frequência LTM · mês a mês (1× … 20× e >20)"
    ws_fm["A1"].font = Font(size=14, bold=True)
    ws_fm["A2"] = (
        "Unique mês = passageiros distintos no mês civil · "
        "Customers = boardings do mês · Unique LTM / faixas = janela rolling até 12m"
    )
    fm_headers = [
        "Mês", "Unique mês", "Customers", "Unique LTM", "≥2", "≥4", "≥2 %", "≥4 %",
    ] + FREQ_LABELS
    for i, h in enumerate(fm_headers, 1):
        ws_fm.cell(4, i, h)
    for r_i, r in enumerate(data["monthly"], 5):
        vals = [
            r["month"],
            r["unique_passengers"],
            r["boardings"],
            r["ltm_unique_customers"],
            r["ltm_ge2"],
            r["ltm_ge4"],
            r["ltm_ge2_pct"],
            r["ltm_ge4_pct"],
        ] + [r.get(k) for k in FREQ_KEYS]
        for c, v in enumerate(vals, 1):
            ws_fm.cell(r_i, c, v if v is not None else None)

    ws = wb.create_sheet("Resumo")
    ws["A1"] = "REVO · Customer growth MoM / YoY"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = (
        f"Fonte {data['source']} · base {data.get('base_start') or data['data_start']} "
        f"→ {data['data_end']} · {data['months_available']} meses"
    )
    cards = [
        ("Unique customers", s.get("unique_customers_all_time")),
        ("Customers (boardings)", s.get("total_boardings")),
        ("Cumulativo unique", s.get("cumulative_unique_end")),
        ("Unique LTM", s.get("ltm_unique_customers")),
        ("MoM cumulativo %", s.get("latest_mom_cumulative_pct")),
        ("YoY cumulativo %", s.get("latest_yoy_cumulative_pct")),
        ("Missões (Sigtrip)", s.get("total_missions") or s.get("total_flights")),
        ("Pernas (legs)", s.get("total_flight_legs")),
        ("Missões LTM", s.get("ltm_missions")),
        ("LTM ≥2", s.get("ltm_ge2")),
        ("LTM ≥4", s.get("ltm_ge4")),
    ]
    for i, (lab, val) in enumerate(cards):
        ws.cell(4, 1 + i, lab)
        ws.cell(5, 1 + i, val if val is not None else "—")

    ws_m = wb.create_sheet("Mensal")
    headers = [
        "Mês", "Novos", "Cumulativo unique", "Unique mês", "Customers", "Missões",
        "MoM cumulativo %", "YoY cumulativo %",
        "LTM unique", "LTM ≥2", "Δ ≥2 vs −12m", "LTM ≥4", "Δ ≥4 vs −12m",
        "≥2 %", "≥4 %",
    ] + [f"LTM {lab}" for lab in FREQ_LABELS]
    for i, h in enumerate(headers, 1):
        ws_m.cell(1, i, h)
    for r_i, r in enumerate(data["monthly"], 2):
        vals = [
            r["month"], r["new_customers"], r["cumulative_unique_customers"],
            r["unique_passengers"], r["boardings"], r["flights"],
            r["mom_cumulative_pct"], r["yoy_cumulative_pct"],
            r["ltm_unique_customers"], r["ltm_ge2"], r.get("ltm_ge2_delta_vs_12m"),
            r["ltm_ge4"], r.get("ltm_ge4_delta_vs_12m"),
            r["ltm_ge2_pct"], r["ltm_ge4_pct"],
        ] + [r.get(k) for k in FREQ_KEYS]
        for c, v in enumerate(vals, 1):
            ws_m.cell(r_i, c, v if v is not None else None)
    last = 1 + len(data["monthly"])

    def add_line(title, min_col, anchor):
        ch = LineChart()
        ch.title = title
        ch.height = 10
        ch.width = 16
        ch.add_data(Reference(ws_m, min_col=min_col, min_row=1, max_row=last), titles_from_data=True)
        ch.set_categories(Reference(ws_m, min_col=1, min_row=2, max_row=last))
        ws_m.add_chart(ch, anchor)

    add_line("Cumulativo unique", 3, "V2")
    add_line("MoM cumulativo %", 7, "V20")
    add_line("YoY cumulativo %", 8, "V38")
    add_line("LTM ≥2", 10, "V56")
    add_line("LTM ≥4", 12, "V74")

    ws_r = wb.create_sheet("Top Rotas")
    for i, h in enumerate(["Rota", "Boardings", "Flights"], 1):
        ws_r.cell(1, i, h)
    for r_i, r in enumerate(data.get("top_routes") or [], 2):
        ws_r.cell(r_i, 1, r.get("route"))
        ws_r.cell(r_i, 2, r.get("boardings"))
        ws_r.cell(r_i, 3, r.get("flights"))

    ws_p = wb.create_sheet("Top Passageiros")
    for i, h in enumerate(["Nome", "Identity", "Boardings", "Datas", "Primeiro", "Último"], 1):
        ws_p.cell(1, i, h)
    for r_i, r in enumerate(data.get("top_passengers") or [], 2):
        ws_p.cell(r_i, 1, r.get("name"))
        ws_p.cell(r_i, 2, r.get("identity_key"))
        ws_p.cell(r_i, 3, r.get("boardings"))
        ws_p.cell(r_i, 4, r.get("distinct_dates"))
        ws_p.cell(r_i, 5, r.get("first_in_window"))
        ws_p.cell(r_i, 6, r.get("last_in_window"))

    wb.save(path)


def main() -> None:
    env = load_env()
    base = env.get("MANIFESTS_API_BASE") or "https://web-production-9b4c2.up.railway.app"
    api_key = env.get("API_KEY") or ""

    try:
        csv_text, summary, monthly_api, routes, top = fetch_api(base, api_key)
        (OUT / "boardings_api.csv").write_text(csv_text, encoding="utf-8")
        source = f"api:{base}"
        print(f"Fetched API {base}")
    except Exception as exc:
        print(f"API fetch failed ({exc}); using local boardings_api.csv")
        csv_text = (OUT / "boardings_api.csv").read_text(encoding="utf-8")
        summary = json.loads((OUT / "summary.json").read_text())
        monthly_api = json.loads((OUT / "monthly.json").read_text())
        routes = json.loads((OUT / "routes.json").read_text()) if (OUT / "routes.json").exists() else []
        top = json.loads((OUT / "top_passengers.json").read_text()) if (OUT / "top_passengers.json").exists() else []
        source = f"local-csv+cache ({base})"

    rows = drop_cancelled_rows(list(csv.DictReader(StringIO(csv_text))))
    # Persist a clean local CSV (no cancelled tabs)
    if rows:
        buf = StringIO()
        w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()), lineterminator="\n")
        w.writeheader()
        w.writerows(rows)
        (OUT / "boardings_api.csv").write_text(buf.getvalue(), encoding="utf-8")

    data = compute(rows, summary, monthly_api, routes, top, source)
    sf_kpis = load_salesforce_kpis()
    if sf_kpis:
        data["salesforce"] = sf_kpis
        print("Included Salesforce KPIs from salesforce_kpis.json")
    else:
        print("No salesforce_kpis.json — run pull_salesforce_kpis.py to add hours/revenue")

    (OUT / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    (OUT / "monthly.json").write_text(json.dumps(monthly_api, indent=2), encoding="utf-8")
    (OUT / "data.js").write_text(
        "window.KPI_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    (OUT / "index.html").write_text(HTML, encoding="utf-8")
    write_excel(data, OUT / "revo-customer-kpis.xlsx")

    print("Dashboard:", OUT / "index.html")
    print("Excel:", OUT / "revo-customer-kpis.xlsx")
    print(json.dumps(data["summary"], indent=2))
    print("months", data["months_available"], data["data_start"], "→", data["data_end"])
    print(
        "Recurrence LTM",
        {
            "month": data["summary"].get("latest_month"),
            "unique": data["summary"].get("ltm_unique_customers"),
            "ge2": data["summary"].get("ltm_ge2"),
            "ge2_delta": data["summary"].get("ltm_ge2_delta_vs_12m"),
            "ge4": data["summary"].get("ltm_ge4"),
            "ge4_delta": data["summary"].get("ltm_ge4_delta_vs_12m"),
            "freq": {
                row["label"]: row["count"]
                for row in (data["summary"].get("ltm_freq_rows") or [])
            },
            "period_freq": {
                row["label"]: row["count"]
                for row in ((data.get("period_frequency") or {}).get("freq_rows") or [])
            },
        },
    )
    print(
        "Snapshots",
        [
            {
                "label": sn.get("label"),
                "ge2": sn.get("ltm_ge2"),
                "d2": sn.get("ltm_ge2_delta_vs_12m"),
                "ge4": sn.get("ltm_ge4"),
                "d4": sn.get("ltm_ge4_delta_vs_12m"),
                "gap": sn.get("data_gap"),
            }
            for sn in data.get("snapshots") or []
        ],
    )
    print(
        "MoM cumul",
        [(r["month"], r["mom_cumulative_pct"]) for r in data["monthly"] if r["mom_cumulative_pct"] is not None][-6:],
    )
    print(
        "YoY cumul",
        [(r["month"], r["yoy_cumulative_pct"]) for r in data["monthly"] if r["yoy_cumulative_pct"] is not None],
    )


if __name__ == "__main__":
    main()
