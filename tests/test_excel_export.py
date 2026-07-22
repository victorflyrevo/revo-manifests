"""Excel workbook export for customer KPIs."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.db import Base
from app.excel_export import build_customer_kpi_workbook
from app.models import Boarding, Flight, Passenger


def _session() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


def test_workbook_has_sheets_and_charts():
    db = _session()
    p1 = Passenger(identity_key="doc:a", display_name="Alice", total_boardings=0)
    p2 = Passenger(identity_key="doc:b", display_name="Bob", total_boardings=0)
    db.add_all([p1, p2])
    db.flush()

    for i, (who, d) in enumerate(
        [
            (p1, date(2025, 1, 10)),
            (p1, date(2025, 3, 10)),
            (p2, date(2025, 6, 10)),
        ],
        start=1,
    ):
        fl = Flight(
            fingerprint=f"fp-{i}",
            source_file="t.xlsx",
            sheet_name="x",
            flight_date=d,
            origin_code="SJK",
            dest_code="RAO",
            pax_count=1,
        )
        db.add(fl)
        db.flush()
        db.add(
            Boarding(
                flight_id=fl.id,
                passenger_id=who.id,
                flight_date=d,
                passenger_name_raw=who.display_name,
                origin_code="SJK",
                dest_code="RAO",
            )
        )
        who.total_boardings = (who.total_boardings or 0) + 1
    db.commit()

    raw = build_customer_kpi_workbook(db, months=12)
    assert raw[:2] == b"PK"  # zip/xlsx signature

    wb = load_workbook(BytesIO(raw))
    assert "Resumo" in wb.sheetnames
    assert "KPIs Mensais" in wb.sheetnames
    assert "Operacional Mensal" in wb.sheetnames
    assert "Top Rotas" in wb.sheetnames
    assert "Top Passageiros" in wb.sheetnames

    ws = wb["KPIs Mensais"]
    assert ws["A3"].value == "Mês"
    assert len(ws._charts) >= 2
