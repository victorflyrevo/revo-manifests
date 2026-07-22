"""Parse REVO manifesto Excel workbooks and CSV bases (one flight group = one flight)."""

from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from typing import Any, BinaryIO, Iterable, Optional

import openpyxl

from app.corrections import (
    extract_sheet_ddmm,
    is_skippable_sheet,
    resolve_flight_date,
    year_hint_from_filename as _year_hint_from_filename,
)

SKIP_PREFIXES = ("base de dados", "xxxx")

# Flat "Base de Dados" / boardings CSV column aliases (normalized header → field)
_CSV_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "flight_date": (
        "flight_date",
        "data",
        "date",
        "data_voo",
        "data do voo",
        "dt",
    ),
    "flight_time": (
        "flight_time",
        "hora",
        "time",
        "horario",
        "horário",
        "hora_voo",
    ),
    "origin": ("origin", "origem", "from", "dep", "partida"),
    "destination": ("destination", "destino", "to", "arr", "chegada"),
    "origin_code": ("origin_code", "origem_code", "origem_iata", "dep_code"),
    "dest_code": ("dest_code", "destino_code", "destino_iata", "arr_code"),
    "passenger_name": (
        "passenger_name",
        "nome",
        "name",
        "nome_passageiro",
        "nome passageiro",
        "passageiro",
        "pax",
    ),
    "document": (
        "document",
        "documento",
        "doc",
        "cpf",
        "rg",
        "passaporte",
        "documento_passageiro",
    ),
    "aircraft_reg": (
        "aircraft_reg",
        "matricula",
        "matrícula",
        "aircraft",
        "aeronave",
        "prefixo",
        "reg",
    ),
    "aircraft_code": ("aircraft_code", "codigo_aeronave", "ac", "tail_code"),
    "sheet_name": ("sheet_name", "aba", "sheet", "voo", "flight", "manifesto"),
    "source_file": ("source_file", "arquivo", "fonte"),
}


@dataclass
class ParsedPassenger:
    name: str
    document: Optional[str]
    identity_key: str
    document_normalized: Optional[str]


@dataclass
class ParsedFlight:
    sheet_name: str
    flight_date: Optional[date]
    flight_time: Optional[str]
    origin: Optional[str]
    destination: Optional[str]
    origin_code: Optional[str]
    dest_code: Optional[str]
    aircraft_reg: Optional[str]
    aircraft_code: Optional[str]
    passengers: list[ParsedPassenger] = field(default_factory=list)
    fingerprint: str = ""


@dataclass
class ParseResult:
    source_file: str
    flights: list[ParsedFlight]
    skipped_sheets: int = 0


def norm_name(s: Any) -> Optional[str]:
    if s is None:
        return None
    text = re.sub(r"\s+", " ", str(s).strip())
    if not text:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.upper()


def norm_doc(s: Any) -> Optional[str]:
    """Normalize passenger document to a stable canonical form (prefers CPF)."""
    from app.identity import canonical_document

    return canonical_document(s)


# Must fit flights.origin_code / dest_code (VARCHAR(8))
AIRPORT_CODE_MAX = 8


def airport_code(s: Any) -> Optional[str]:
    """Extract a short airport/helipad code from a free-text origin/destination.

    Examples:
      "SDXQ - International Plaza" → "SDXQ"
      "ZCAPS - Joaquim Egídio" → "ZCAPS"
      "SBSP" → "SBSP"
    """
    if s is None:
        return None
    text = str(s).strip()
    if not text:
        return None
    upper = text.upper()
    # ICAO (4), local helipads (4–5), and similar leading tokens up to DB width
    m = re.match(r"^([A-Z0-9]{3,8})\b", upper)
    if m:
        return m.group(1)[:AIRPORT_CODE_MAX]
    # Fallback: first token before separators, never longer than the column
    token = re.split(r"[\s\-–—|/]+", upper, maxsplit=1)[0]
    token = re.sub(r"[^A-Z0-9]", "", token)
    if len(token) >= 3:
        return token[:AIRPORT_CODE_MAX]
    return None


def identity_key(name: str, document: Any) -> tuple[str, Optional[str]]:
    """Document-preferred identity; CPF variants collapse to the same key."""
    from app.identity import identity_key as _identity_key

    return _identity_key(name, document)


def sheet_ddmm(sn: str) -> Optional[tuple[int, int]]:
    """Backward-compatible wrapper around corrections.extract_sheet_ddmm."""
    return extract_sheet_ddmm(sn)


def year_hint_from_filename(filename: str) -> Optional[int]:
    """Backward-compatible wrapper around corrections.year_hint_from_filename."""
    return _year_hint_from_filename(filename)


def resolve_date(
    cell_date: Any, sheet_name: str, filename: str
) -> Optional[date]:
    """Backward-compatible wrapper around corrections.resolve_flight_date."""
    return resolve_flight_date(cell_date, sheet_name, filename)


def flight_fingerprint(fl: ParsedFlight, source_file: str) -> str:
    pax = "|".join(sorted(p.identity_key for p in fl.passengers))
    raw = "|".join(
        [
            fl.flight_date.isoformat() if fl.flight_date else "",
            fl.origin_code or "",
            fl.dest_code or "",
            fl.flight_time or "",
            fl.aircraft_code or fl.aircraft_reg or "",
            pax,
            # keep sheet canon lightly to distinguish parallel empty flights
            re.sub(r"^C[oó]pia de\s+", "", fl.sheet_name.strip(), flags=re.I),
            source_file,
        ]
    )
    # Cross-file dedup: same operational identity without source_file
    operational = "|".join(
        [
            fl.flight_date.isoformat() if fl.flight_date else "",
            fl.origin_code or "",
            fl.dest_code or "",
            fl.flight_time or "",
            fl.aircraft_code or fl.aircraft_reg or "",
            pax,
        ]
    )
    return hashlib.sha256(operational.encode("utf-8")).hexdigest()


def _parse_sheet_rows(rows: Iterable[Any], sn: str, filename: str) -> ParsedFlight:
    """Parse one sheet from an iterable of row tuples — no passenger cap."""
    flight_date = origin = dest = hora = matricula = None
    passengers: list[ParsedPassenger] = []
    pax_mode = False
    empty_pax_streak = 0

    for row in rows:
        if not row:
            if pax_mode:
                empty_pax_streak += 1
                if empty_pax_streak >= 5:
                    break
            continue

        # Normalize to a sequence we can index
        row = tuple(row)
        if any(isinstance(v, str) and "TOTAL" in v.upper() for v in row if v is not None):
            break

        label = row[0] if len(row) > 0 else None
        if not pax_mode and label is not None:
            lab = str(label).strip().lower()
            val = row[3] if len(row) > 3 else None
            if lab.startswith("data"):
                flight_date = val
            elif lab.startswith("hora"):
                hora = val
            elif lab.startswith("origem"):
                origin = val
            elif lab.startswith("destino"):
                dest = val
            elif "matr" in lab:
                matricula = row[13] if len(row) > 13 else val
            elif label == "#" or (
                len(row) > 1
                and isinstance(row[1], str)
                and "nome" in str(row[1]).lower()
            ):
                pax_mode = True
            continue

        if not pax_mode:
            # Header row detected by column B containing "Nome Passageiro"
            if len(row) > 1 and isinstance(row[1], str) and "nome" in row[1].lower():
                pax_mode = True
            continue

        name = row[1] if len(row) > 1 else None
        doc = row[7] if len(row) > 7 else None
        if name is None or str(name).strip() == "":
            empty_pax_streak += 1
            if empty_pax_streak >= 5:
                break
            continue

        empty_pax_streak = 0
        name_s = str(name).strip()
        key, doc_n = identity_key(name_s, doc)
        passengers.append(
            ParsedPassenger(
                name=name_s,
                document=str(doc).strip() if doc is not None else None,
                identity_key=key,
                document_normalized=doc_n,
            )
        )

    # String dates (common in ODS / CSV-exported sheets)
    if not isinstance(flight_date, (date, datetime)):
        flight_date = _parse_csv_date(flight_date)

    fd = resolve_date(flight_date, sn, filename)

    hora_s = None
    if isinstance(hora, datetime):
        hora_s = hora.strftime("%H:%M")
    elif hasattr(hora, "hour"):
        hora_s = f"{hora.hour:02d}:{hora.minute:02d}"
    elif hora is not None:
        hora_s = _parse_csv_time(hora) or str(hora)[:16]

    ac = None
    m = re.search(r"\b(OOE\d*|OMB\d*|OMH\d*)\b", sn.upper())
    if m:
        ac = m.group(1)

    fl = ParsedFlight(
        sheet_name=sn,
        flight_date=fd,
        flight_time=hora_s,
        origin=str(origin).strip() if origin else None,
        destination=str(dest).strip() if dest else None,
        origin_code=airport_code(origin),
        dest_code=airport_code(dest),
        aircraft_reg=str(matricula).strip() if matricula else None,
        aircraft_code=ac,
        passengers=passengers,
    )
    fl.fingerprint = flight_fingerprint(fl, filename)
    return fl


def _parse_sheet(ws: Any, sn: str, filename: str) -> ParsedFlight:
    """Parse one Excel sheet via openpyxl."""
    return _parse_sheet_rows(
        ws.iter_rows(min_row=1, max_col=16, values_only=True), sn, filename
    )


def parse_workbook(file_obj: BinaryIO, filename: str) -> ParseResult:
    """Parse every sheet in the workbook — no sheet-count limit."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    flights: list[ParsedFlight] = []
    skipped = 0

    for sn in wb.sheetnames:
        if is_skippable_sheet(sn) or sn.strip().lower().startswith(SKIP_PREFIXES):
            skipped += 1
            continue
        flights.append(_parse_sheet(wb[sn], sn, filename))

    wb.close()
    return ParseResult(source_file=filename, flights=flights, skipped_sheets=skipped)


def _ods_cell_value(cell: Any) -> Any:
    """Extract a Python value from an odfpy TableCell."""
    from odf.text import P

    vtype = (cell.getAttribute("valuetype") or "").lower()
    if vtype == "float":
        raw = cell.getAttribute("value")
        if raw is None or raw == "":
            return None
        try:
            num = float(raw)
            return int(num) if num.is_integer() else num
        except ValueError:
            return raw
    if vtype == "date":
        raw = cell.getAttribute("datevalue") or cell.getAttribute("value")
        if not raw:
            return None
        try:
            return datetime.fromisoformat(str(raw)[:19]).date()
        except ValueError:
            return _parse_csv_date(raw)
    if vtype == "time":
        raw = cell.getAttribute("timevalue") or cell.getAttribute("value")
        if not raw:
            return None
        # ODS time: PThhHmmMssS
        m = re.match(
            r"^PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(?:\.\d+)?)S)?$",
            str(raw).upper(),
        )
        if m:
            hh = int(m.group(1) or 0)
            mm = int(m.group(2) or 0)
            return f"{hh:02d}:{mm:02d}"
        return str(raw)
    if vtype == "boolean":
        raw = (cell.getAttribute("booleanvalue") or "").lower()
        return raw == "true"

    parts: list[str] = []
    for p in cell.getElementsByType(P):
        if p.firstChild is not None and getattr(p.firstChild, "data", None):
            parts.append(p.firstChild.data)
    text = "".join(parts).strip()
    if text:
        return text
    stringvalue = cell.getAttribute("stringvalue")
    return stringvalue if stringvalue not in (None, "") else None


def _ods_table_rows(table: Any, max_col: int = 16) -> list[tuple[Any, ...]]:
    from odf.table import TableCell, TableRow

    rows: list[tuple[Any, ...]] = []
    for tr in table.getElementsByType(TableRow):
        values: list[Any] = []
        for cell in tr.getElementsByType(TableCell):
            if len(values) >= max_col:
                break
            val = _ods_cell_value(cell)
            repeat = cell.getAttribute("numbercolumnsrepeated")
            n = 1
            if repeat:
                try:
                    n = max(1, int(repeat))
                except ValueError:
                    n = 1
            # Huge repeats are used for trailing empty columns — cap them
            n = min(n, max_col - len(values))
            values.extend([val] * n)
        if len(values) < max_col:
            values.extend([None] * (max_col - len(values)))
        rows.append(tuple(values[:max_col]))
    return rows


def parse_ods(file_obj: BinaryIO, filename: str) -> ParseResult:
    """Parse every sheet in an OpenDocument Spreadsheet (.ods)."""
    from odf.opendocument import load
    from odf.table import Table

    # odfpy load() wants a path or file-like with read/seek
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except Exception:  # noqa: BLE001
            pass
    doc = load(file_obj)
    flights: list[ParsedFlight] = []
    skipped = 0

    for table in doc.spreadsheet.getElementsByType(Table):
        sn = table.getAttribute("name") or "Sheet"
        if is_skippable_sheet(sn) or sn.strip().lower().startswith(SKIP_PREFIXES):
            skipped += 1
            continue
        rows = _ods_table_rows(table)
        flights.append(_parse_sheet_rows(rows, sn, filename))

    return ParseResult(source_file=filename, flights=flights, skipped_sheets=skipped)


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[\s_]+", " ", text)
    return text.strip()


def _header_map(headers: Iterable[Any]) -> dict[str, int]:
    """Map logical field name → column index from a header row."""
    alias_to_field: dict[str, str] = {}
    for field_name, aliases in _CSV_FIELD_ALIASES.items():
        for alias in aliases:
            alias_to_field[_norm_header(alias)] = field_name

    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = _norm_header(header)
        if not key:
            continue
        field_name = alias_to_field.get(key)
        if field_name and field_name not in mapping:
            mapping[field_name] = idx
    return mapping


def _decode_csv_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _detect_delimiter(sample: str) -> str:
    lines = [ln for ln in sample.splitlines() if ln.strip()]
    head = "\n".join(lines[:5])
    if not head:
        return ","
    try:
        dialect = csv.Sniffer().sniff(head, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        pass
    # Brazilian Excel exports often use ';'
    counts = {d: head.count(d) for d in (",", ";", "\t", "|")}
    return max(counts, key=counts.get) or ","


def _parse_csv_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    # Excel serial date (rare in CSV, but harmless)
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            serial = float(text)
            if 20000 <= serial <= 80000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
        except ValueError:
            pass

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def _parse_csv_time(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{value.hour:02d}:{value.minute:02d}"

    text = str(value).strip()
    if not text:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if m:
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    return text[:16]


def _cell(row: list[Any], idx: Optional[int]) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _looks_like_flat_header(headers: list[Any]) -> bool:
    mapping = _header_map(headers)
    return "passenger_name" in mapping and (
        "flight_date" in mapping or "origin" in mapping or "destination" in mapping
    )


def _parse_flat_csv_rows(
    rows: list[list[Any]], headers: list[Any], filename: str
) -> list[ParsedFlight]:
    mapping = _header_map(headers)
    name_idx = mapping.get("passenger_name")
    if name_idx is None:
        raise ValueError("CSV base missing passenger name column (nome / passenger_name)")

    groups: "OrderedDict[tuple, ParsedFlight]" = OrderedDict()

    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        name_raw = _cell(row, name_idx)
        if name_raw is None or str(name_raw).strip() == "":
            continue
        # Skip repeated header-like rows mid-file
        if "nome" in str(name_raw).strip().lower() and len(str(name_raw).split()) <= 3:
            continue

        name_s = str(name_raw).strip()
        doc = _cell(row, mapping.get("document"))
        key, doc_n = identity_key(name_s, doc)
        pax = ParsedPassenger(
            name=name_s,
            document=str(doc).strip() if doc is not None and str(doc).strip() else None,
            identity_key=key,
            document_normalized=doc_n,
        )

        origin = _cell(row, mapping.get("origin"))
        dest = _cell(row, mapping.get("destination"))
        origin_s = str(origin).strip() if origin is not None and str(origin).strip() else None
        dest_s = str(dest).strip() if dest is not None and str(dest).strip() else None
        origin_code = airport_code(_cell(row, mapping.get("origin_code")) or origin_s)
        dest_code = airport_code(_cell(row, mapping.get("dest_code")) or dest_s)

        fd = _parse_csv_date(_cell(row, mapping.get("flight_date")))
        hora_s = _parse_csv_time(_cell(row, mapping.get("flight_time")))

        matricula = _cell(row, mapping.get("aircraft_reg"))
        ac_raw = _cell(row, mapping.get("aircraft_code"))
        aircraft_reg = (
            str(matricula).strip() if matricula is not None and str(matricula).strip() else None
        )
        aircraft_code = (
            str(ac_raw).strip().upper()
            if ac_raw is not None and str(ac_raw).strip()
            else None
        )

        sheet_raw = _cell(row, mapping.get("sheet_name"))
        if sheet_raw is not None and str(sheet_raw).strip():
            sheet_name = str(sheet_raw).strip()
        else:
            bits = [
                fd.isoformat() if fd else "nodate",
                f"{origin_code or '?'}-{dest_code or '?'}",
                hora_s or "notime",
                aircraft_reg or aircraft_code or "",
            ]
            sheet_name = "CSV " + " ".join(b for b in bits if b)

        # Apply the same date corrections used for Excel/ODS sheets
        fd = resolve_date(fd, sheet_name, filename)

        if is_skippable_sheet(sheet_name):
            continue

        if not aircraft_code:
            m = re.search(r"\b(OOE\d*|OMB\d*|OMH\d*)\b", sheet_name.upper())
            if m:
                aircraft_code = m.group(1)

        group_key = (
            fd.isoformat() if fd else "",
            hora_s or "",
            origin_code or "",
            dest_code or "",
            aircraft_reg or "",
            aircraft_code or "",
            sheet_name,
        )

        if group_key not in groups:
            fl = ParsedFlight(
                sheet_name=sheet_name,
                flight_date=fd,
                flight_time=hora_s,
                origin=origin_s,
                destination=dest_s,
                origin_code=origin_code,
                dest_code=dest_code,
                aircraft_reg=aircraft_reg,
                aircraft_code=aircraft_code,
                passengers=[],
            )
            groups[group_key] = fl
        groups[group_key].passengers.append(pax)

    flights: list[ParsedFlight] = []
    for fl in groups.values():
        fl.fingerprint = flight_fingerprint(fl, filename)
        flights.append(fl)
    return flights


def _parse_manifest_style_csv(
    rows: list[list[Any]], filename: str, sheet_name: str
) -> ParsedFlight:
    """Parse a CSV that mirrors a single manifesto sheet layout."""
    flight_date = origin = dest = hora = matricula = None
    passengers: list[ParsedPassenger] = []
    pax_mode = False
    empty_pax_streak = 0

    for row in rows:
        # Normalize to at least 16 logical columns like Excel sheets
        cells = list(row) + [None] * max(0, 16 - len(row))
        if not any(c is not None and str(c).strip() != "" for c in cells):
            if pax_mode:
                empty_pax_streak += 1
                if empty_pax_streak >= 5:
                    break
            continue

        if any(isinstance(v, str) and "TOTAL" in v.upper() for v in cells if v is not None):
            break

        label = cells[0]
        if not pax_mode and label is not None:
            lab = str(label).strip().lower()
            # Prefer column D (index 3), else first non-empty after label
            val = cells[3]
            if val is None or str(val).strip() == "":
                for c in cells[1:]:
                    if c is not None and str(c).strip() != "":
                        val = c
                        break
            if lab.startswith("data"):
                flight_date = val
            elif lab.startswith("hora"):
                hora = val
            elif lab.startswith("origem"):
                origin = val
            elif lab.startswith("destino"):
                dest = val
            elif "matr" in lab:
                matricula = cells[13] if cells[13] is not None else val
            elif label == "#" or (
                isinstance(cells[1], str) and "nome" in str(cells[1]).lower()
            ):
                pax_mode = True
            continue

        if not pax_mode:
            if isinstance(cells[1], str) and "nome" in cells[1].lower():
                pax_mode = True
            continue

        name = cells[1]
        doc = cells[7]
        if name is None or str(name).strip() == "":
            empty_pax_streak += 1
            if empty_pax_streak >= 5:
                break
            continue

        empty_pax_streak = 0
        name_s = str(name).strip()
        key, doc_n = identity_key(name_s, doc)
        passengers.append(
            ParsedPassenger(
                name=name_s,
                document=str(doc).strip() if doc is not None else None,
                identity_key=key,
                document_normalized=doc_n,
            )
        )

    fd = resolve_date(_parse_csv_date(flight_date), sheet_name, filename)
    hora_s = _parse_csv_time(hora)

    ac = None
    m = re.search(r"\b(OOE\d*|OMB\d*|OMH\d*)\b", sheet_name.upper())
    if m:
        ac = m.group(1)

    fl = ParsedFlight(
        sheet_name=sheet_name,
        flight_date=fd,
        flight_time=hora_s,
        origin=str(origin).strip() if origin else None,
        destination=str(dest).strip() if dest else None,
        origin_code=airport_code(origin),
        dest_code=airport_code(dest),
        aircraft_reg=str(matricula).strip() if matricula else None,
        aircraft_code=ac,
        passengers=passengers,
    )
    fl.fingerprint = flight_fingerprint(fl, filename)
    return fl


def parse_csv(file_obj: BinaryIO, filename: str) -> ParseResult:
    """
    Parse a CSV manifesto / Base de Dados export.

    Supports:
    - Flat tabular base (one row per passenger/boarding) with PT/EN headers
    - Single-sheet manifesto layout (Data/Hora/Origem… + passenger list)
    """
    raw = file_obj.read()
    if isinstance(raw, str):
        text = raw
    else:
        text = _decode_csv_bytes(raw)

    if not text.strip():
        raise ValueError("Empty CSV file")

    delimiter = _detect_delimiter(text)
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = [list(r) for r in reader]
    if not rows:
        raise ValueError("Empty CSV file")

    # Drop fully empty trailing rows
    while rows and all(not str(c).strip() for c in rows[-1]):
        rows.pop()

    stem = _file_stem(filename)
    header = rows[0]
    if _looks_like_flat_header(header):
        flights = _parse_flat_csv_rows(rows[1:], header, filename)
        if not flights:
            raise ValueError("CSV base has headers but no passenger rows")
        return ParseResult(source_file=filename, flights=flights, skipped_sheets=0)

    # Manifest-style: whole file is one flight
    fl = _parse_manifest_style_csv(rows, filename, stem or "CSV")
    if not fl.passengers and not fl.flight_date and not fl.origin:
        raise ValueError(
            "Unrecognized CSV format. Expected a Base de Dados table "
            "(columns like Data, Origem, Destino, Nome, Documento) "
            "or a manifesto-style sheet export."
        )
    return ParseResult(source_file=filename, flights=[fl], skipped_sheets=0)


def _file_stem(filename: str) -> str:
    name = filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    if "." in name:
        return name.rsplit(".", 1)[0]
    return name


def parse_bytes(data: bytes, filename: str) -> ParseResult:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(BytesIO(data), filename)
    if lower.endswith(".ods"):
        return parse_ods(BytesIO(data), filename)
    return parse_workbook(BytesIO(data), filename)


def content_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()
