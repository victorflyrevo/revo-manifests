"""Build an Excel workbook with customer KPIs, supporting tables, and charts."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.customer_kpis import compute_customer_kpis
from app.models import Boarding, Flight, Passenger

HEADER_FILL = PatternFill("solid", fgColor="1A211C")
HEADER_FONT = Font(color="E8EEE6", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1A211C")
SUB_FONT = Font(name="Calibri", size=10, color="666666")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", size=18, bold=True, color="1A211C")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ACCENT_FILL = PatternFill("solid", fgColor="F7F1E3")


def _autosize(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def _write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
        r += 1
    return r - 1


def _monthly_trends(db: Session) -> list[dict]:
    dialect = db.bind.dialect.name if db.bind else "sqlite"
    if dialect == "postgresql":
        month_expr = func.to_char(Boarding.flight_date, "YYYY-MM")
    else:
        month_expr = func.strftime("%Y-%m", Boarding.flight_date)

    rows = db.execute(
        select(
            month_expr.label("month"),
            func.count(Boarding.id).label("boardings"),
            func.count(func.distinct(Boarding.flight_id)).label("flights"),
            func.count(func.distinct(Boarding.passenger_id)).label("unique_passengers"),
        )
        .where(Boarding.flight_date.is_not(None))
        .group_by(month_expr)
        .order_by(month_expr)
    ).all()
    return [
        {
            "month": r.month,
            "boardings": r.boardings,
            "flights": r.flights,
            "unique_passengers": r.unique_passengers,
        }
        for r in rows
    ]


def _top_routes(db: Session, days: int = 365, limit: int = 25) -> list[dict]:
    end = date.today()
    start = end - timedelta(days=days)
    route = func.concat(
        func.coalesce(Boarding.origin_code, "?"),
        "→",
        func.coalesce(Boarding.dest_code, "?"),
    )
    rows = db.execute(
        select(
            route.label("route"),
            func.count(Boarding.id).label("boardings"),
            func.count(func.distinct(Boarding.flight_id)).label("flights"),
        )
        .where(Boarding.flight_date >= start, Boarding.flight_date <= end)
        .group_by(route)
        .order_by(func.count(Boarding.id).desc())
        .limit(limit)
    ).all()
    return [
        {"route": r.route, "boardings": r.boardings, "flights": r.flights} for r in rows
    ]


def _top_passengers(db: Session, days: int = 365, limit: int = 50) -> list[dict]:
    end = date.today()
    start = end - timedelta(days=days)
    rows = db.execute(
        select(
            Passenger.display_name,
            Passenger.identity_key,
            func.count(Boarding.id).label("boardings"),
            func.count(func.distinct(Boarding.flight_date)).label("distinct_dates"),
            func.min(Boarding.flight_date).label("first_in_window"),
            func.max(Boarding.flight_date).label("last_in_window"),
        )
        .join(Boarding, Boarding.passenger_id == Passenger.id)
        .where(Boarding.flight_date >= start, Boarding.flight_date <= end)
        .group_by(Passenger.id, Passenger.display_name, Passenger.identity_key)
        .order_by(func.count(Boarding.id).desc())
        .limit(limit)
    ).all()
    return [
        {
            "name": r.display_name,
            "identity_key": r.identity_key,
            "boardings": r.boardings,
            "distinct_dates": r.distinct_dates,
            "first_in_window": r.first_in_window.isoformat() if r.first_in_window else "",
            "last_in_window": r.last_in_window.isoformat() if r.last_in_window else "",
        }
        for r in rows
    ]


def build_customer_kpi_workbook(db: Session, months: int = 12) -> bytes:
    kpis = compute_customer_kpis(db, months=months)
    monthly_ops = _monthly_trends(db)
    routes = _top_routes(db, days=365, limit=25)
    top_pax = _top_passengers(db, days=365, limit=50)

    wb = Workbook()

    # ---- Resumo ----
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "REVO · Customer KPIs"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        f"Histórico {kpis.get('data_start') or '—'} → {kpis.get('data_end') or '—'} · "
        f"Janela {kpis.get('ltm_start') or '—'} → {kpis.get('anchor_month') or '—'} "
        f"({kpis.get('months_available', 0)} meses)"
    )
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    summary = kpis.get("summary") or {}
    cards = [
        ("Unique customers LTM", summary.get("unique_customers_ltm", 0)),
        ("Repeat rate LTM", f"{summary.get('repeat_rate_pct', 0)}%"),
        ("New customers LTM", summary.get("new_customers_ltm", 0)),
        ("Repeaters LTM", summary.get("repeat_customers_ltm", 0)),
        ("One-time LTM", summary.get("one_time_customers_ltm", 0)),
        ("Cumulative unique (end)", summary.get("cumulative_unique_end", 0)),
    ]
    for i, (label, value) in enumerate(cards):
        col = 1 + (i % 3) * 2
        row = 4 + (i // 3) * 3
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.font = KPI_LABEL_FONT
        label_cell.fill = ACCENT_FILL
        value_cell = ws.cell(row=row + 1, column=col, value=value)
        value_cell.font = KPI_VALUE_FONT
        value_cell.fill = ACCENT_FILL
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(
            start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
        )

    ws["A11"] = "Como ler"
    ws["A11"].font = Font(bold=True, size=12)
    ws["A12"] = (
        "• Cumulative unique: crescimento acumulado de clientes que voaram pela "
        "primeira vez dentro da janela exibida."
    )
    ws["A13"] = (
        "• Repeat rate: % dos clientes únicos na janela móvel de 12 meses "
        "(terminando em cada mês) que voaram ≥2 vezes nessa janela."
    )
    ws["A14"] = (
        "• A base usa o histórico mais longo disponível no banco "
        "(data_start → data_end)."
    )
    ws["A16"] = "Abas"
    ws["A16"].font = Font(bold=True, size=12)
    ws["A17"] = "KPIs Mensais — série + gráficos prontos (cumulative + repeat rate)"
    ws["A18"] = "Operacional Mensal — boardings / flights / uniques por mês"
    ws["A19"] = "Top Rotas — últimas 365 dias"
    ws["A20"] = "Top Passageiros — últimas 365 dias"
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 28

    # ---- KPIs Mensais + charts ----
    ws_m = wb.create_sheet("KPIs Mensais")
    ws_m["A1"] = "Customer KPIs mensais (LTM)"
    ws_m["A1"].font = TITLE_FONT
    headers = [
        "Mês",
        "Novos clientes",
        "Cumulativo unique",
        "LTM unique",
        "LTM repeaters",
        "LTM one-time",
        "Repeat rate %",
        "Window start",
        "Window end",
    ]
    rows = [
        [
            r["month"],
            r["new_customers"],
            r["cumulative_unique_customers"],
            r["ltm_unique_customers"],
            r["ltm_repeat_customers"],
            r["ltm_one_time_customers"],
            r["repeat_rate_pct"],
            r["window_start"],
            r["window_end"],
        ]
        for r in kpis.get("monthly") or []
    ]
    last_data_row = _write_table(ws_m, 3, headers, rows)
    _autosize(ws_m)

    if rows:
        # Cumulative line chart
        chart1 = LineChart()
        chart1.title = "Crescimento cumulativo de unique customers"
        chart1.style = 10
        chart1.y_axis.title = "Clientes"
        chart1.x_axis.title = "Mês"
        chart1.height = 10
        chart1.width = 18
        data1 = Reference(ws_m, min_col=3, min_row=3, max_row=last_data_row)
        cats1 = Reference(ws_m, min_col=1, min_row=4, max_row=last_data_row)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        if chart1.series:
            chart1.series[0].graphicalProperties.line.solidFill = "C4A35A"
        ws_m.add_chart(chart1, "A" + str(last_data_row + 3))

        # Repeat rate line chart
        chart2 = LineChart()
        chart2.title = "Repeat rate mês a mês (LTM rolling)"
        chart2.style = 12
        chart2.y_axis.title = "%"
        chart2.x_axis.title = "Mês"
        chart2.y_axis.scaling.min = 0
        chart2.y_axis.scaling.max = 100
        chart2.height = 10
        chart2.width = 18
        data2 = Reference(ws_m, min_col=7, min_row=3, max_row=last_data_row)
        cats2 = Reference(ws_m, min_col=1, min_row=4, max_row=last_data_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        if chart2.series:
            chart2.series[0].graphicalProperties.line.solidFill = "7CB89A"
            chart2.series[0].dLbls = DataLabelList()
            chart2.series[0].dLbls.showVal = True
        ws_m.add_chart(chart2, "J" + str(last_data_row + 3))

        # New vs repeaters combo-style second block
        chart3 = LineChart()
        chart3.title = "Novos clientes vs LTM repeaters"
        chart3.style = 10
        chart3.y_axis.title = "Clientes"
        chart3.height = 10
        chart3.width = 18
        data3 = Reference(ws_m, min_col=2, min_row=3, max_col=2, max_row=last_data_row)
        data3b = Reference(ws_m, min_col=5, min_row=3, max_col=5, max_row=last_data_row)
        cats3 = Reference(ws_m, min_col=1, min_row=4, max_row=last_data_row)
        chart3.add_data(data3, titles_from_data=True)
        chart3.add_data(data3b, titles_from_data=True)
        chart3.set_categories(cats3)
        ws_m.add_chart(chart3, "A" + str(last_data_row + 22))

    # ---- Operacional mensal ----
    ws_o = wb.create_sheet("Operacional Mensal")
    ws_o["A1"] = "Boardings / flights / uniques por mês (todo o histórico)"
    ws_o["A1"].font = TITLE_FONT
    op_headers = ["Mês", "Boardings", "Flights", "Unique passengers"]
    op_rows = [
        [r["month"], r["boardings"], r["flights"], r["unique_passengers"]]
        for r in monthly_ops
    ]
    op_last = _write_table(ws_o, 3, op_headers, op_rows)
    _autosize(ws_o)
    if op_rows:
        chart_op = LineChart()
        chart_op.title = "Boardings e unique passengers por mês"
        chart_op.style = 10
        chart_op.height = 10
        chart_op.width = 18
        data_op = Reference(ws_o, min_col=2, min_row=3, max_col=4, max_row=op_last)
        cats_op = Reference(ws_o, min_col=1, min_row=4, max_row=op_last)
        chart_op.add_data(data_op, titles_from_data=True)
        chart_op.set_categories(cats_op)
        ws_o.add_chart(chart_op, "F3")

    # ---- Top rotas ----
    ws_r = wb.create_sheet("Top Rotas")
    ws_r["A1"] = "Top rotas · últimos 365 dias"
    ws_r["A1"].font = TITLE_FONT
    _write_table(
        ws_r,
        3,
        ["Rota", "Boardings", "Flights"],
        [[r["route"], r["boardings"], r["flights"]] for r in routes],
    )
    _autosize(ws_r)

    # ---- Top passageiros ----
    ws_p = wb.create_sheet("Top Passageiros")
    ws_p["A1"] = "Top passageiros · últimos 365 dias"
    ws_p["A1"].font = TITLE_FONT
    _write_table(
        ws_p,
        3,
        [
            "Nome",
            "Identity key",
            "Boardings",
            "Datas distintas",
            "Primeiro na janela",
            "Último na janela",
        ],
        [
            [
                r["name"],
                r["identity_key"],
                r["boardings"],
                r["distinct_dates"],
                r["first_in_window"],
                r["last_in_window"],
            ]
            for r in top_pax
        ],
    )
    _autosize(ws_p)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
